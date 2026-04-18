# Copyright (C) 2026 Project Librarian contributors
#
# This file is part of Project Librarian.
#
# Project Librarian is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# Project Librarian is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Project Librarian. If not, see <https://www.gnu.org/licenses/>.

"""Excel and CSV keyword indexing helpers."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xlsb", ".xls", ".csv"}


def _record_skip(
    skipped_files: list[dict[str, str]] | None,
    file_path: Path,
    stage: str,
    reason: str,
) -> None:
    """Append a normalized skipped-file record."""
    if skipped_files is None:
        return
    skipped_files.append({"path": file_path.as_posix(), "stage": stage, "reason": reason})


def list_excel_files(folder_path: Path) -> list[Path]:
    """Return known spreadsheet files beneath folder_path."""
    if not folder_path.exists():
        return []
    return sorted(path for path in folder_path.rglob("*") if path.suffix.lower() in EXCEL_SUFFIXES and path.is_file())


def discover_headers(file_path: Path, skipped_files: list[dict[str, str]] | None = None) -> list[str]:
    """Return first-row header values for a spreadsheet-like file."""
    if file_path.suffix.lower() == ".csv":
        try:
            with file_path.open("r", encoding="utf-8", newline="") as handle:
                reader = csv.reader(handle)
                first_row = next(reader, [])
                return [str(value).strip() for value in first_row if str(value).strip()]
        except (OSError, UnicodeDecodeError, csv.Error) as exc:
            _record_skip(skipped_files, file_path=file_path, stage="excel_headers", reason=f"csv_error:{exc.__class__.__name__}")
            return []

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1), []):
            text = str(cell.value or "").strip()
            if text:
                headers.append(text)
        return headers
    except Exception as exc:
        _record_skip(skipped_files, file_path=file_path, stage="excel_headers", reason=f"xlsx_error:{exc.__class__.__name__}")
        return []


def index_excel_rows(
    folder_path: Path,
    keyword_columns: list[str],
    limit: int = 200,
    skipped_files: list[dict[str, str]] | None = None,
) -> list[dict[str, object]]:
    """Collect keyword-matching row fields from spreadsheets for search integration."""
    records: list[dict[str, object]] = []
    selected = {column.strip().lower() for column in keyword_columns if column.strip()}
    if not selected:
        return records

    for file_path in list_excel_files(folder_path):
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            try:
                with file_path.open("r", encoding="utf-8", newline="") as handle:
                    reader = csv.DictReader(handle)
                    for row_number, row in enumerate(reader, start=2):
                        for key, value in row.items():
                            if str(key).strip().lower() not in selected:
                                continue
                            text = str(value or "").strip()
                            if not text:
                                continue
                            records.append(
                                {
                                    "file": file_path.as_posix(),
                                    "sheet": "csv",
                                    "row": row_number,
                                    "field": str(key),
                                    "value": text,
                                }
                            )
                            if len(records) >= limit:
                                return records
            except (OSError, UnicodeDecodeError, csv.Error) as exc:
                _record_skip(skipped_files, file_path=file_path, stage="excel_rows", reason=f"csv_error:{exc.__class__.__name__}")
                continue
            continue

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            _record_skip(skipped_files, file_path=file_path, stage="excel_rows", reason=f"xlsx_error:{exc.__class__.__name__}")
            continue
        for sheet in workbook.worksheets:
            header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1), []))
            headers = [str(cell.value or "").strip() for cell in header_cells]
            wanted_indexes = [idx for idx, header in enumerate(headers) if header.lower() in selected]
            if not wanted_indexes:
                continue

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                for index in wanted_indexes:
                    value = str(row[index] or "").strip() if index < len(row) else ""
                    if not value:
                        continue
                    records.append(
                        {
                            "file": file_path.as_posix(),
                            "sheet": sheet.title,
                            "row": row_number,
                            "field": headers[index],
                            "value": value,
                        }
                    )
                    if len(records) >= limit:
                        return records
    return records
