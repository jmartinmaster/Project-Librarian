import argparse
import csv
import hashlib
import html
import json
import os
import re
import secrets
import shlex
import shutil
import subprocess
import threading
import time
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path

from symbol_index import DEFAULT_OUTPUT_DIR as SYMBOL_INDEX_OUTPUT_DIR, JSON_OUTPUT_NAME as SYMBOL_INDEX_JSON_NAME, SymbolIndexError, generate_symbol_index


DEFAULT_OUTPUT_DIR = Path("build") / "project-librarian"
SNAPSHOT_NAME = "librarian-snapshot.json"
HISTORY_NAME = "change-history.jsonl"
CORPUS_NAME = "search-corpus.json"
EXCEL_LIBRARY_CONFIG_NAME = "excel_library_config.json"
EXCEL_EXTENSIONS = (".xlsx", ".xls", ".xlsm", ".xlsb", ".csv")
DEFAULT_EXCEL_KEYWORD_COLUMNS = ["Downtime Code", "Shop Order", "Part Number", "Date"]
DRAFTS_DIR_NAME = "drafts"
AI_CONTEXT_DIR_NAME = "ai-context"
SNAPSHOT_VERSION = 2
DEFAULT_AI_MODEL = "qwen2.5-coder:14b"
DEFAULT_MCP_TRANSPORT = "streamable-http"
DEFAULT_MCP_HOST = "127.0.0.1"
DEFAULT_MCP_PORT = 8765
DEFAULT_REFRESH_INTERVAL_SECONDS = 30.0
DEFAULT_LIBRARY_WATCH_INTERVAL_SECONDS = 5.0
HTTP_AUTH_ENV_NAME = "PROJECT_LIBRARIAN_TOKEN"
HTTP_AUTH_COOKIE_NAME = "project_librarian_token"
HTTP_AUTH_HEADER_NAME = "X-Project-Librarian-Token"
README_TARGET_NAME = "README.md"
CHANGELOG_TARGET_NAME = "CHANGELOG.md"
DOC_BLOCK_START = "<!-- project-librarian:docs:start -->"
DOC_BLOCK_END = "<!-- project-librarian:docs:end -->"
SEARCHABLE_EXTENSIONS = {
    ".code-workspace",
    ".desktop",
    ".json",
    ".md",
    ".py",
    ".sh",
    ".spec",
    ".toml",
    ".txt",
    ".yaml",
    ".yml",
}
EXCLUDED_DIRECTORY_NAMES = {
    ".git",
    ".idea",
    ".venv",
    ".vscode",
    "__pycache__",
    "build",
    "data",
    "dist",
    "env",
    "exports",
    "logs",
    "venv",
}
TOKEN_PATTERN = re.compile(r"[A-Za-z0-9_./:-]+")
STATUS_TOKEN_PATTERN = re.compile(r"[A-Za-z?]+")
DIFF_HUNK_PATTERN = re.compile(r"@@ -\d+(?:,\d+)? \+(\d+)(?:,(\d+))? @@")
CHANGELOG_HEADING_PATTERN = re.compile(r"^## \[(?P<version>[^\]]+)\] - .+$", re.MULTILINE)
AREA_ORDER = ("app", "controllers", "models", "views", "docs", "scripts", "root")
DOC_SUGGESTIONS_BY_AREA = {
    "app": ["README.md", "CHANGELOG.md", "docs/help/"],
    "controllers": ["README.md", "CHANGELOG.md", "docs/help/"],
    "models": ["README.md", "CHANGELOG.md", "docs/production_log_json_architecture.md"],
    "views": ["README.md", "CHANGELOG.md", "docs/help/"],
    "docs": ["README.md", "CHANGELOG.md"],
    "scripts": ["README.md", "docs/ai-delegation/"],
    "root": ["README.md", "CHANGELOG.md"],
}


class ProjectLibrarianError(RuntimeError):
    pass


@dataclass
class LibrarianWorkspace:
    repo_root: Path
    output_dir: Path
    snapshot: dict
    corpus: dict
    history: list

    @classmethod
    def load(cls, repo_root=None, output_dir=None, refresh_if_missing=True, refresh_first=False):
        resolved_repo_root = Path(repo_root or _repo_root_from_here()).resolve()
        resolved_output_dir = _resolve_output_dir(resolved_repo_root, output_dir)
        if refresh_first:
            build_librarian_snapshot(repo_root=resolved_repo_root, output_dir=resolved_output_dir)
        snapshot = _load_snapshot(
            repo_root=resolved_repo_root,
            output_dir=resolved_output_dir,
            refresh_if_missing=refresh_if_missing,
        )
        corpus = _load_corpus(
            repo_root=resolved_repo_root,
            output_dir=resolved_output_dir,
            refresh_if_missing=refresh_if_missing,
        )
        history = _load_history(
            repo_root=resolved_repo_root,
            output_dir=resolved_output_dir,
        )
        return cls(
            repo_root=resolved_repo_root,
            output_dir=resolved_output_dir,
            snapshot=snapshot,
            corpus=corpus,
            history=history,
        )

    def refresh(self):
        result = build_librarian_snapshot(repo_root=self.repo_root, output_dir=self.output_dir)
        self.snapshot = result["snapshot"]
        self.corpus = _load_corpus(repo_root=self.repo_root, output_dir=self.output_dir, refresh_if_missing=False)
        self.history = _load_history(repo_root=self.repo_root, output_dir=self.output_dir)
        return result

    @property
    def changed_files(self):
        return self.snapshot.get("git", {}).get("changed_files", [])

    @property
    def changed_paths(self):
        return {item.get("path") for item in self.changed_files if item.get("path")}

    @property
    def file_records(self):
        return self.snapshot.get("files", [])

    @property
    def symbol_records(self):
        return self.snapshot.get("symbols", [])

    @property
    def file_lookup(self):
        return {record.get("path"): record for record in self.file_records if record.get("path")}


class LibrarianService:
    def __init__(
        self,
        repo_root=None,
        output_dir=None,
        refresh_if_missing=True,
        refresh_first=False,
        refresh_interval_seconds=DEFAULT_REFRESH_INTERVAL_SECONDS,
        library_watch_interval_seconds=DEFAULT_LIBRARY_WATCH_INTERVAL_SECONDS,
        start_refresh_worker=True,
    ):
        self.repo_root = Path(repo_root or _repo_root_from_here()).resolve()
        self.output_dir = _resolve_output_dir(self.repo_root, output_dir)
        self.refresh_interval_seconds = max(0.0, float(refresh_interval_seconds or 0.0))
        self._lock = threading.RLock()
        self._stop_event = threading.Event()
        self._refresh_thread = None
        self._library_watch_thread = None
        self._library_watch_signature = ""
        self._library_watch_file_count = 0
        self._library_watch_last_scan_at = None
        self._library_watch_last_change_at = None
        self._library_watch_last_refresh_at = None
        self._library_watch_error = None
        self.library_watch_interval_seconds = max(1.0, float(library_watch_interval_seconds or DEFAULT_LIBRARY_WATCH_INTERVAL_SECONDS))
        self._last_refresh_at = None
        self._last_refresh_error = None
        self._workspace = LibrarianWorkspace.load(
            repo_root=self.repo_root,
            output_dir=self.output_dir,
            refresh_if_missing=refresh_if_missing,
            refresh_first=refresh_first,
        )
        self._last_refresh_at = _utc_now_text()
        self.configure_library_watcher(force_restart=True)
        if start_refresh_worker and self.refresh_interval_seconds > 0:
            self.start_refresh_worker()

    def start_refresh_worker(self):
        if self._refresh_thread is not None and self._refresh_thread.is_alive():
            return
        self._stop_event.clear()
        self._refresh_thread = threading.Thread(
            target=self._refresh_loop,
            name="project-librarian-refresh",
            daemon=True,
        )
        self._refresh_thread.start()

    def stop(self, join_timeout=2.0):
        self._stop_event.set()
        if self._refresh_thread is not None and self._refresh_thread.is_alive():
            self._refresh_thread.join(timeout=max(0.1, float(join_timeout or 0.1)))
        if self._library_watch_thread is not None and self._library_watch_thread.is_alive():
            self._library_watch_thread.join(timeout=max(0.1, float(join_timeout or 0.1)))

    def _refresh_loop(self):
        while not self._stop_event.wait(self.refresh_interval_seconds):
            try:
                self.refresh()
            except Exception as exc:
                self._last_refresh_error = str(exc)

    def _compute_library_watch_signature(self):
        hasher = hashlib.sha1()
        file_count = 0
        for path in sorted((candidate for candidate in self.repo_root.rglob("*") if candidate.is_file() and candidate.suffix.lower() in SEARCHABLE_EXTENSIONS and all(part not in candidate.parts for part in EXCLUDED_DIRECTORY_NAMES)), key=lambda value: str(value.relative_to(self.repo_root)).lower()):
            try:
                stat = path.stat()
            except OSError:
                continue
            rel_path = str(path.relative_to(self.repo_root)).replace("\\", "/").lower()
            hasher.update(rel_path.encode("utf-8", errors="ignore"))
            hasher.update(str(int(getattr(stat, "st_mtime_ns", int(stat.st_mtime * 1_000_000_000)))).encode("ascii"))
            hasher.update(str(int(stat.st_size)).encode("ascii"))
            file_count += 1
        return hasher.hexdigest(), file_count

    def _library_watch_loop(self):
        while not self._stop_event.wait(self.library_watch_interval_seconds):
            try:
                signature, file_count = self._compute_library_watch_signature()
                self._library_watch_last_scan_at = _utc_now_text()
                self._library_watch_file_count = file_count
                if not self._library_watch_signature:
                    self._library_watch_signature = signature
                    self._library_watch_error = None
                    continue
                if signature != self._library_watch_signature:
                    self._library_watch_signature = signature
                    self._library_watch_last_change_at = _utc_now_text()
                    self.refresh()
                    self._library_watch_last_refresh_at = self._last_refresh_at
                self._library_watch_error = None
            except Exception as exc:
                self._library_watch_error = str(exc)

    def configure_library_watcher(self, force_restart=False):
        thread_running = bool(self._library_watch_thread and self._library_watch_thread.is_alive())
        if not force_restart and thread_running:
            return
        self._library_watch_signature = ""
        self._library_watch_file_count = 0
        self._library_watch_last_scan_at = None
        self._library_watch_last_change_at = None
        self._library_watch_last_refresh_at = None
        self._library_watch_error = None
        if thread_running:
            return
        self._stop_event.clear()
        self._library_watch_thread = threading.Thread(
            target=self._library_watch_loop,
            name="project-librarian-library-watch",
            daemon=True,
        )
        self._library_watch_thread.start()

    def library_watch_payload(self):
        return {
            "enabled": True,
            "watcher_running": bool(self._library_watch_thread and self._library_watch_thread.is_alive()),
            "interval_seconds": self.library_watch_interval_seconds,
            "last_scan_at": self._library_watch_last_scan_at,
            "last_change_at": self._library_watch_last_change_at,
            "last_refresh_at": self._library_watch_last_refresh_at,
            "tracked_files": self._library_watch_file_count,
            "last_error": self._library_watch_error,
        }

    def _with_workspace(self, callback):
        with self._lock:
            return callback(self._workspace)

    def refresh(self):
        def _refresh(workspace):
            result = workspace.refresh()
            self._last_refresh_at = _utc_now_text()
            self._last_refresh_error = None
            return result

        return self._with_workspace(_refresh)

    def status_payload(self):
        def _build(workspace):
            summary = workspace.snapshot.get("summary", {})
            git_snapshot = workspace.snapshot.get("git", {})
            return {
                "repo_root": str(self.repo_root),
                "output_dir": str(self.output_dir),
                "branch": git_snapshot.get("branch", "unknown"),
                "files": summary.get("files", 0),
                "symbols": summary.get("symbols", 0),
                "changed_files": summary.get("changed_files", 0),
                "history_entries": summary.get("history_entries", len(workspace.history)),
                "refresh_interval_seconds": self.refresh_interval_seconds,
                "refresh_worker_running": bool(self._refresh_thread and self._refresh_thread.is_alive()),
                "last_refresh_at": self._last_refresh_at,
                "last_refresh_error": self._last_refresh_error,
                "library_watcher": self.library_watch_payload(),
            }

        return self._with_workspace(_build)

    def stats_text(self):
        return self._with_workspace(format_workspace_stats)

    def snapshot_payload(self):
        return self._with_workspace(lambda workspace: workspace.snapshot)

    def history_payload(self, limit=10):
        def _build(workspace):
            return {
                "count": min(len(workspace.history), max(1, int(limit))),
                "history": workspace.history[-max(1, int(limit)):],
                "formatted": format_history_report(workspace.history, limit=limit),
            }

        return self._with_workspace(_build)

    def search_payload(self, query, scope="all", limit=20, area=None, changed_only=False, path_filter=None):
        def _build(workspace):
            results = search_snapshot(
                workspace.snapshot,
                workspace.corpus,
                query,
                scope=scope,
                limit=limit,
                area=area,
                changed_only=changed_only,
                path_filter=path_filter,
            )
            return {
                "count": len(results),
                "results": results,
                "formatted": format_search_results(results),
            }

        return self._with_workspace(_build)

    def changes_payload(self, limit=20, status_filter=None, area=None, path_filter=None, include_commits=True):
        def _build(workspace):
            filtered = _filter_changed_files(workspace.snapshot, status_filter=status_filter, area=area, path_filter=path_filter)
            return {
                "count": len(filtered),
                "changes": filtered[: max(1, int(limit))],
                "formatted": format_change_report(
                    workspace.snapshot,
                    limit=limit,
                    status_filter=status_filter,
                    area=area,
                    path_filter=path_filter,
                    include_commits=include_commits,
                ),
            }

        return self._with_workspace(_build)

    def show_excerpt(self, path_text, query=None, line=None, context=3):
        return self._with_workspace(lambda workspace: show_file_excerpt(workspace, path_text, query=query, line=line, context=context))

    def docs_draft_payload(self, title=None, changed_only=True, output_path=None, apply=False, target_path=README_TARGET_NAME):
        def _build(workspace):
            content = generate_docs_draft(workspace, title=title, changed_only=changed_only)
            payload = {
                "content": content,
                "changed_only": changed_only,
            }
            if apply:
                resolved_path = apply_docs_update(self.repo_root, content, target_path=target_path)
                payload["applied_path"] = str(resolved_path)
            else:
                resolved_path = _write_generated_output(
                    workspace.output_dir,
                    DRAFTS_DIR_NAME,
                    "docs_draft",
                    content,
                    output_path=output_path,
                    output_base_dir=self.repo_root,
                )
                payload["output_path"] = str(resolved_path)
            return payload

        return self._with_workspace(_build)

    def changelog_draft_payload(self, version_text=None, release_date=None, changed_only=True, output_path=None, apply=False, target_path=CHANGELOG_TARGET_NAME):
        def _build(workspace):
            content = generate_changelog_draft(
                workspace,
                version_text=version_text,
                release_date=release_date,
                changed_only=changed_only,
            )
            payload = {
                "content": content,
                "changed_only": changed_only,
                "version": version_text or "Unreleased",
            }
            if apply:
                resolved_path = apply_changelog_update(
                    self.repo_root,
                    content,
                    version_label=version_text or "Unreleased",
                    target_path=target_path,
                )
                payload["applied_path"] = str(resolved_path)
            else:
                resolved_path = _write_generated_output(
                    workspace.output_dir,
                    DRAFTS_DIR_NAME,
                    "changelog_draft",
                    content,
                    output_path=output_path,
                    output_base_dir=self.repo_root,
                )
                payload["output_path"] = str(resolved_path)
            return payload

        return self._with_workspace(_build)

    def ai_models_payload(self, preferred_model=DEFAULT_AI_MODEL, ollama_host=None):
        status = collect_ai_runtime_status(self.repo_root, preferred_model=preferred_model, ollama_host=ollama_host)
        return {
            "status": status,
            "formatted": format_ai_model_list(status),
        }

    def ai_doctor_payload(self, preferred_model=DEFAULT_AI_MODEL, ollama_host=None):
        status = collect_ai_runtime_status(self.repo_root, preferred_model=preferred_model, ollama_host=ollama_host)
        return {
            "status": status,
            "formatted": format_ai_status_report(status),
        }

    def ai_summary_payload(self, task, mode="analysis", model=DEFAULT_AI_MODEL, changed_only=True, ollama_host=None):
        def _build(workspace):
            result = run_ai_summary(
                workspace,
                task=task,
                mode=mode,
                model=model,
                changed_only=changed_only,
                ollama_host=ollama_host,
            )
            return result

        return self._with_workspace(_build)


def _repo_root_from_here():
    return Path(__file__).resolve().parent


def _resolve_output_dir(repo_root, output_dir=None):
    resolved_repo_root = Path(repo_root or _repo_root_from_here()).resolve()
    candidate = Path(output_dir) if output_dir is not None else DEFAULT_OUTPUT_DIR
    if candidate.is_absolute():
        return candidate.resolve()
    return (resolved_repo_root / candidate).resolve()


def _utc_now_text():
    return datetime.now(timezone.utc).isoformat()


def _today_text():
    return datetime.now(timezone.utc).date().isoformat()


def _relative_path_text(path, repo_root):
    return path.relative_to(repo_root).as_posix()


def _file_area(relative_path):
    if relative_path.startswith("docs/"):
        return "docs"
    if relative_path.startswith("app/controllers/"):
        return "controllers"
    if relative_path.startswith("app/models/"):
        return "models"
    if relative_path.startswith("app/views/"):
        return "views"
    if relative_path.startswith("app/"):
        return "app"
    if relative_path.startswith("scripts/"):
        return "scripts"
    return "root"


def _iter_searchable_files(repo_root):
    files = []
    for root_path, dir_names, file_names in os.walk(repo_root, topdown=True):
        dir_names[:] = [
            dir_name
            for dir_name in dir_names
            if dir_name not in EXCLUDED_DIRECTORY_NAMES and not dir_name.startswith(".venv")
        ]
        for file_name in file_names:
            suffix = Path(file_name).suffix.lower()
            if suffix not in SEARCHABLE_EXTENSIONS:
                continue
            files.append(Path(root_path) / file_name)
    return sorted(files, key=lambda path: _relative_path_text(path, repo_root))


def _extract_title(relative_path, text):
    if relative_path.endswith(".md"):
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("#"):
                return stripped.lstrip("#").strip()
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            return stripped[:120]
    return relative_path


def _token_count(text):
    return len(TOKEN_PATTERN.findall(text))


def _preview_for_query(text, query, limit=160):
    lowered_query = str(query or "").lower()
    for line_number, line in enumerate(text.splitlines(), start=1):
        if lowered_query in line.lower():
            snippet = line.strip()
            if len(snippet) > limit:
                snippet = f"{snippet[: limit - 3]}..."
            return {"line": line_number, "text": snippet}
    preview_line = next((line.strip() for line in text.splitlines() if line.strip()), "")
    if len(preview_line) > limit:
        preview_line = f"{preview_line[: limit - 3]}..."
    return {"line": 1 if preview_line else None, "text": preview_line}


def _run_git_command(repo_root, *args):
    completed = subprocess.run(
        ["git", *args],
        cwd=repo_root,
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        return ""
    return completed.stdout.strip()


def _run_shell_command(command, cwd=None, env=None):
    return subprocess.run(
        command,
        cwd=cwd,
        check=False,
        capture_output=True,
        text=True,
        env=env,
    )


def _run_git_status(repo_root):
    completed = subprocess.run(
        ["git", "status", "--porcelain=v1", "-z", "--untracked-files=all"],
        cwd=repo_root,
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        return []

    entries = completed.stdout.split("\0")
    changed_files = []
    index = 0
    while index < len(entries):
        entry = entries[index]
        index += 1
        if not entry:
            continue
        raw_status = entry[:2]
        status = raw_status.strip() or "??"
        path_text = entry[3:]
        change_record = {
            "status": status,
            "xy": raw_status,
            "path": path_text,
            "area": _file_area(path_text),
        }
        if "R" in status or "C" in status:
            source_path = entries[index] if index < len(entries) else ""
            if source_path:
                change_record["source_path"] = source_path
                index += 1
        changed_files.append(change_record)
    return changed_files


def _collect_recent_commits(repo_root, limit=5):
    completed = subprocess.run(
        [
            "git",
            "log",
            f"--max-count={max(1, int(limit))}",
            "--date=short",
            "--pretty=format:%H%x1f%h%x1f%ad%x1f%an%x1f%s",
        ],
        cwd=repo_root,
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0 or not completed.stdout.strip():
        return []

    commits = []
    for line in completed.stdout.splitlines():
        parts = line.split("\x1f")
        if len(parts) != 5:
            continue
        commits.append(
            {
                "commit": parts[0],
                "short_commit": parts[1],
                "date": parts[2],
                "author": parts[3],
                "subject": parts[4],
            }
        )
    return commits


def _collect_git_status_payload(repo_root, commit_limit=12):
    branch_name = _run_git_command(repo_root, "rev-parse", "--abbrev-ref", "HEAD") or "unknown"
    upstream_name = _run_git_command(repo_root, "rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{upstream}")
    ahead_count = 0
    behind_count = 0
    if upstream_name:
        counts_result = _run_shell_command(
            ["git", "rev-list", "--left-right", "--count", f"{upstream_name}...HEAD"],
            cwd=repo_root,
        )
        if counts_result.returncode == 0:
            parts = counts_result.stdout.strip().split()
            if len(parts) >= 2:
                behind_count = _coerce_int(parts[0], 0, minimum=0, maximum=999999)
                ahead_count = _coerce_int(parts[1], 0, minimum=0, maximum=999999)

    changed_files = _run_git_status(repo_root)
    return {
        "branch": branch_name,
        "upstream": upstream_name,
        "ahead": ahead_count,
        "behind": behind_count,
        "changed_files": changed_files,
        "changed_count": len(changed_files),
        "status_counts": _counts_from_items(changed_files, "status"),
        "area_counts": _counts_from_items(changed_files, "area"),
        "recent_commits": _collect_recent_commits(repo_root, limit=commit_limit),
    }


def _run_git_action(repo_root, args):
    completed = _run_shell_command(["git", *args], cwd=repo_root)
    stdout_text = (completed.stdout or "").strip()
    stderr_text = (completed.stderr or "").strip()
    if completed.returncode != 0:
        detail = stderr_text or stdout_text or f"git {' '.join(args)} failed"
        raise ProjectLibrarianError(detail)
    return {
        "ok": True,
        "stdout": stdout_text,
        "stderr": stderr_text,
        "command": ["git", *args],
    }


def _read_git_diff_text(repo_root, path_text=None, staged=False):
    command = ["git", "diff", "--no-ext-diff", "--color=never"]
    if staged:
        command.append("--staged")
    if path_text:
        command.extend(["--", path_text])
    completed = _run_shell_command(command, cwd=repo_root)
    if completed.returncode != 0:
        raise ProjectLibrarianError((completed.stderr or completed.stdout or "Unable to read git diff.").strip())
    return completed.stdout or ""


def _read_git_commit_review(repo_root, commit_ref):
    completed = _run_shell_command(
        [
            "git",
            "show",
            "--no-color",
            "--no-ext-diff",
            "--stat",
            "--patch",
            "--find-renames",
            str(commit_ref),
        ],
        cwd=repo_root,
    )
    if completed.returncode != 0:
        raise ProjectLibrarianError((completed.stderr or completed.stdout or "Unable to load commit review.").strip())
    return completed.stdout or ""


def _counts_from_items(items, key_name):
    counts = {}
    for item in items:
        key_value = str(item.get(key_name) or "unknown")
        counts[key_value] = counts.get(key_value, 0) + 1
    return dict(sorted(counts.items(), key=lambda pair: (-pair[1], pair[0])))


def _get_change_record(workspace, path_text):
    for item in workspace.changed_files:
        if item.get("path") == path_text:
            return item
    return {}


def _symbols_for_path(workspace, path_text):
    symbols = [record for record in workspace.symbol_records if record.get("path") == path_text]
    return sorted(symbols, key=lambda record: (record.get("line") or 0, record.get("qualified_name") or record.get("name") or ""))


def _parse_changed_line_numbers(diff_text):
    line_numbers = set()
    for line in str(diff_text or "").splitlines():
        match = DIFF_HUNK_PATTERN.match(line)
        if not match:
            continue
        start_line = max(1, int(match.group(1) or 1))
        line_count = int(match.group(2) or "1")
        if line_count <= 0:
            line_numbers.add(start_line)
            continue
        for line_number in range(start_line, start_line + line_count):
            line_numbers.add(line_number)
    return sorted(line_numbers)


def _collect_changed_line_numbers(repo_root, path_text):
    completed = _run_shell_command(["git", "diff", "--unified=0", "--", path_text], cwd=repo_root)
    if completed.returncode != 0:
        return []
    return _parse_changed_line_numbers(completed.stdout)


def _nearest_symbols(symbols, changed_lines, limit=4):
    if not symbols or not changed_lines:
        return []

    scored_symbols = []
    for symbol in symbols:
        try:
            symbol_line = int(symbol.get("line") or 0)
        except (TypeError, ValueError):
            symbol_line = 0
        if symbol_line <= 0:
            continue
        distance = min(abs(symbol_line - changed_line) for changed_line in changed_lines)
        scored_symbols.append((distance, symbol_line, symbol.get("qualified_name") or symbol.get("name") or "", symbol))

    scored_symbols.sort(key=lambda item: (item[0], item[1], item[2]))
    return [item[3] for item in scored_symbols[: max(1, int(limit))]]


def _collect_touched_symbols(workspace, records):
    touched_symbols = {}
    for record in records:
        path_text = record.get("path")
        if not path_text:
            continue
        symbols = _symbols_for_path(workspace, path_text)
        if not symbols:
            continue

        change_record = _get_change_record(workspace, path_text)
        status_text = str(change_record.get("status") or "")
        if "?" in status_text or "A" in status_text:
            touched_symbols[path_text] = symbols[:8]
            continue

        changed_lines = _collect_changed_line_numbers(workspace.repo_root, path_text)
        if not changed_lines:
            touched_symbols[path_text] = symbols[:4]
            continue

        matches = []
        for symbol in symbols:
            try:
                symbol_line = int(symbol.get("line") or 0)
            except (TypeError, ValueError):
                symbol_line = 0
            if symbol_line <= 0:
                continue
            if any(abs(symbol_line - changed_line) <= 3 for changed_line in changed_lines):
                matches.append(symbol)

        touched_symbols[path_text] = matches[:8] if matches else _nearest_symbols(symbols, changed_lines)
    return touched_symbols


def _format_symbol_label(symbol_record):
    symbol_name = symbol_record.get("qualified_name") or symbol_record.get("name") or "(unknown)"
    symbol_kind = symbol_record.get("kind")
    if symbol_kind:
        return f"{symbol_name} ({symbol_kind})"
    return str(symbol_name)


def _summarize_symbol_labels(symbol_records, limit=4):
    labels = []
    seen = set()
    for symbol_record in symbol_records:
        label = _format_symbol_label(symbol_record)
        if label in seen:
            continue
        labels.append(label)
        seen.add(label)
        if len(labels) >= max(1, int(limit)):
            break
    remaining = max(0, len({ _format_symbol_label(symbol_record) for symbol_record in symbol_records }) - len(labels))
    if remaining > 0:
        labels.append(f"and {remaining} more")
    return ", ".join(labels)


def _collect_area_symbol_summary(records, touched_symbols):
    area_symbols = {}
    for record in records:
        path_text = record.get("path")
        area_name = record.get("area") or _file_area(path_text or "")
        symbols = touched_symbols.get(path_text, [])
        if not symbols:
            continue
        area_symbols.setdefault(area_name, []).extend(symbols)
    return area_symbols


def _recent_commit_subjects(workspace, limit=5):
    return [
        str(commit.get("subject") or "").strip()
        for commit in workspace.snapshot.get("git", {}).get("recent_commits", [])[: max(1, int(limit))]
        if str(commit.get("subject") or "").strip()
    ]


def _collect_git_snapshot(repo_root):
    branch_name = _run_git_command(repo_root, "rev-parse", "--abbrev-ref", "HEAD") or "unknown"
    changed_files = _run_git_status(repo_root)
    return {
        "branch": branch_name,
        "changed_files": changed_files,
        "changed_count": len(changed_files),
        "status_counts": _counts_from_items(changed_files, "status"),
        "area_counts": _counts_from_items(changed_files, "area"),
        "recent_commits": _collect_recent_commits(repo_root, limit=5),
    }


def _load_symbol_payload(repo_root):
    symbol_output_dir = (repo_root / SYMBOL_INDEX_OUTPUT_DIR).resolve()
    generate_symbol_index(repo_root=repo_root, output_dir=symbol_output_dir)
    symbol_json_path = symbol_output_dir / SYMBOL_INDEX_JSON_NAME
    try:
        return json.loads(symbol_json_path.read_text(encoding="utf-8")), symbol_json_path
    except OSError as exc:
        raise ProjectLibrarianError(f"Unable to read symbol index at {symbol_json_path}: {exc}") from exc
    except json.JSONDecodeError as exc:
        raise ProjectLibrarianError(f"Unable to parse symbol index at {symbol_json_path}: {exc}") from exc


def _flatten_symbol_payload(symbol_payload):
    symbol_records = []
    for file_entry in symbol_payload.get("files", []):
        file_path = file_entry.get("path", "")
        module_name = file_entry.get("module_name")

        for variable_entry in file_entry.get("variables", []):
            symbol_records.append(
                {
                    "path": file_path,
                    "line": variable_entry.get("line"),
                    "kind": variable_entry.get("kind"),
                    "name": variable_entry.get("name"),
                    "qualified_name": variable_entry.get("name"),
                    "signature": variable_entry.get("name"),
                    "doc_summary": None,
                    "context": module_name,
                }
            )

        for function_entry in file_entry.get("functions", []):
            symbol_records.append(
                {
                    "path": file_path,
                    "line": function_entry.get("line"),
                    "kind": function_entry.get("kind"),
                    "name": function_entry.get("name"),
                    "qualified_name": function_entry.get("name"),
                    "signature": function_entry.get("signature"),
                    "doc_summary": function_entry.get("doc_summary"),
                    "context": module_name,
                }
            )

        for class_entry in file_entry.get("classes", []):
            class_name = class_entry.get("name")
            symbol_records.append(
                {
                    "path": file_path,
                    "line": class_entry.get("line"),
                    "kind": class_entry.get("kind"),
                    "name": class_name,
                    "qualified_name": class_name,
                    "signature": class_name,
                    "doc_summary": class_entry.get("doc_summary"),
                    "context": module_name,
                }
            )
            for attribute_entry in class_entry.get("attributes", []):
                symbol_records.append(
                    {
                        "path": file_path,
                        "line": attribute_entry.get("line"),
                        "kind": attribute_entry.get("kind"),
                        "name": attribute_entry.get("name"),
                        "qualified_name": f"{class_name}.{attribute_entry.get('name')}",
                        "signature": attribute_entry.get("name"),
                        "doc_summary": None,
                        "context": class_name,
                    }
                )
            for method_entry in class_entry.get("methods", []):
                symbol_records.append(
                    {
                        "path": file_path,
                        "line": method_entry.get("line"),
                        "kind": method_entry.get("kind"),
                        "name": method_entry.get("name"),
                        "qualified_name": f"{class_name}.{method_entry.get('name')}",
                        "signature": method_entry.get("signature"),
                        "doc_summary": method_entry.get("doc_summary"),
                        "context": class_name,
                    }
                )
    return symbol_records


def _build_file_records(repo_root):
    file_records = []
    corpus_records = {}
    total_bytes = 0
    for path in _iter_searchable_files(repo_root):
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue
        except OSError as exc:
            raise ProjectLibrarianError(f"Unable to read {path}: {exc}") from exc

        stat_result = path.stat()
        relative_path = _relative_path_text(path, repo_root)
        total_bytes += stat_result.st_size
        corpus_records[relative_path] = text
        file_records.append(
            {
                "path": relative_path,
                "area": _file_area(relative_path),
                "title": _extract_title(relative_path, text),
                "line_count": len(text.splitlines()),
                "size_bytes": stat_result.st_size,
                "modified_at": datetime.fromtimestamp(stat_result.st_mtime, tz=timezone.utc).isoformat(),
                "token_count": _token_count(text),
            }
        )
    return file_records, corpus_records, total_bytes


def _path_score_adjustment(path_text):
    if path_text.startswith("docs/ai-delegation/archive/"):
        return -30
    if path_text.startswith("app/"):
        return 20
    if path_text.startswith("docs/help/") or path_text == "README.md":
        return 15
    if path_text.startswith("docs/"):
        return 5
    return 0


def build_librarian_snapshot(repo_root=None, output_dir=None):
    repo_root = Path(repo_root or _repo_root_from_here()).resolve()
    output_dir = _resolve_output_dir(repo_root, output_dir)

    try:
        symbol_payload, symbol_json_path = _load_symbol_payload(repo_root)
    except SymbolIndexError as exc:
        raise ProjectLibrarianError(f"Unable to refresh symbol index for librarian: {exc}") from exc

    symbol_records = _flatten_symbol_payload(symbol_payload)
    file_records, corpus_records, total_bytes = _build_file_records(repo_root)
    git_snapshot = _collect_git_snapshot(repo_root)

    snapshot = {
        "snapshot_version": SNAPSHOT_VERSION,
        "generated_at": _utc_now_text(),
        "repo_root": str(repo_root),
        "summary": {
            "files": len(file_records),
            "symbols": len(symbol_records),
            "bytes": total_bytes,
            "changed_files": git_snapshot["changed_count"],
            "history_entries": 0,
        },
        "symbol_index_path": str(symbol_json_path),
        "corpus_path": str(output_dir / CORPUS_NAME),
        "git": git_snapshot,
        "files": file_records,
        "symbols": symbol_records,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    snapshot_path = output_dir / SNAPSHOT_NAME
    history_path = output_dir / HISTORY_NAME
    corpus_path = output_dir / CORPUS_NAME
    snapshot_path.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
    corpus_path.write_text(json.dumps(corpus_records), encoding="utf-8")

    history_record = {
        "generated_at": snapshot["generated_at"],
        "branch": git_snapshot["branch"],
        "changed_count": git_snapshot["changed_count"],
        "changed_files": git_snapshot["changed_files"],
        "status_counts": git_snapshot.get("status_counts", {}),
        "area_counts": git_snapshot.get("area_counts", {}),
        "recent_commits": git_snapshot.get("recent_commits", []),
        "summary": snapshot["summary"],
    }
    with history_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(history_record) + "\n")

    history_entries = len(_load_history(repo_root=repo_root, output_dir=output_dir))
    snapshot["summary"]["history_entries"] = history_entries
    snapshot_path.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")

    return {
        "snapshot": snapshot,
        "snapshot_path": snapshot_path,
        "history_path": history_path,
        "corpus_path": corpus_path,
    }


def refresh_librarian_snapshot(repo_root=None, output_dir=None):
    return build_librarian_snapshot(repo_root=repo_root, output_dir=output_dir)


def _load_snapshot(repo_root=None, output_dir=None, refresh_if_missing=False):
    repo_root = Path(repo_root or _repo_root_from_here()).resolve()
    output_dir = _resolve_output_dir(repo_root, output_dir)
    snapshot_path = output_dir / SNAPSHOT_NAME
    if not snapshot_path.exists():
        if refresh_if_missing:
            return build_librarian_snapshot(repo_root=repo_root, output_dir=output_dir)["snapshot"]
        raise ProjectLibrarianError(f"Snapshot not found at {snapshot_path}. Run refresh first.")
    try:
        snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    except OSError as exc:
        raise ProjectLibrarianError(f"Unable to read librarian snapshot: {exc}") from exc
    except json.JSONDecodeError as exc:
        raise ProjectLibrarianError(f"Unable to parse librarian snapshot: {exc}") from exc
    snapshot.setdefault("summary", {})
    snapshot["summary"].setdefault("history_entries", len(_load_history(repo_root=repo_root, output_dir=output_dir)))
    snapshot.setdefault("git", {})
    snapshot["git"].setdefault("changed_files", [])
    snapshot["git"].setdefault("status_counts", _counts_from_items(snapshot["git"].get("changed_files", []), "status"))
    snapshot["git"].setdefault("area_counts", _counts_from_items(snapshot["git"].get("changed_files", []), "area"))
    snapshot["git"].setdefault("recent_commits", [])
    return snapshot


def _load_corpus(repo_root=None, output_dir=None, refresh_if_missing=False):
    repo_root = Path(repo_root or _repo_root_from_here()).resolve()
    output_dir = _resolve_output_dir(repo_root, output_dir)
    corpus_path = output_dir / CORPUS_NAME
    if not corpus_path.exists():
        if refresh_if_missing:
            build_librarian_snapshot(repo_root=repo_root, output_dir=output_dir)
            return _load_corpus(repo_root=repo_root, output_dir=output_dir, refresh_if_missing=False)
        raise ProjectLibrarianError(f"Search corpus not found at {corpus_path}. Run refresh first.")
    try:
        return json.loads(corpus_path.read_text(encoding="utf-8"))
    except OSError as exc:
        raise ProjectLibrarianError(f"Unable to read search corpus: {exc}") from exc
    except json.JSONDecodeError as exc:
        raise ProjectLibrarianError(f"Unable to parse search corpus: {exc}") from exc


def _load_history(repo_root=None, output_dir=None):
    repo_root = Path(repo_root or _repo_root_from_here()).resolve()
    output_dir = _resolve_output_dir(repo_root, output_dir)
    history_path = output_dir / HISTORY_NAME
    if not history_path.exists():
        return []

    records = []
    try:
        with history_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                raw_line = line.strip()
                if not raw_line:
                    continue
                try:
                    records.append(json.loads(raw_line))
                except json.JSONDecodeError:
                    continue
    except OSError as exc:
        raise ProjectLibrarianError(f"Unable to read change history at {history_path}: {exc}") from exc
    return records


def _normalize_status_filter(status_filter):
    if not status_filter:
        return set()
    tokens = STATUS_TOKEN_PATTERN.findall(str(status_filter).upper())
    return {token for token in tokens if token}


def _matches_common_filters(path_text, area=None, path_filter=None, changed_only=False, changed_paths=None):
    candidate_path = str(path_text or "")
    if area and _file_area(candidate_path) != area:
        return False
    if path_filter and path_filter.lower() not in candidate_path.lower():
        return False
    if changed_only and candidate_path not in (changed_paths or set()):
        return False
    return True


def _score_text_record(record, file_text, query_tokens, query_text, changed_paths=None):
    searchable_parts = [record.get("path", ""), record.get("title", ""), file_text]
    searchable_text = "\n".join(searchable_parts).lower()
    if query_text not in searchable_text and not all(token in searchable_text for token in query_tokens):
        return None

    score = 0
    path_text = record.get("path", "").lower()
    title_text = record.get("title", "").lower()
    if query_text in path_text:
        score += 80
    if query_text in title_text:
        score += 50
    if query_text in file_text.lower():
        score += 20
    for token in query_tokens:
        if token in path_text:
            score += 15
        elif token in title_text:
            score += 10
        elif token in searchable_text:
            score += 5
    score += _path_score_adjustment(record.get("path", ""))
    preview = _preview_for_query(file_text, query_text)
    return {
        "type": "file",
        "score": score,
        "path": record.get("path"),
        "line": preview.get("line"),
        "title": record.get("title"),
        "preview": preview.get("text"),
        "area": record.get("area"),
        "changed": record.get("path") in (changed_paths or set()),
    }


def _score_symbol_record(record, query_tokens, query_text, changed_paths=None):
    searchable_text = " ".join(
        str(part or "")
        for part in (
            record.get("name"),
            record.get("qualified_name"),
            record.get("signature"),
            record.get("doc_summary"),
            record.get("path"),
        )
    ).lower()
    if query_text not in searchable_text and not all(token in searchable_text for token in query_tokens):
        return None

    score = 0
    if query_text in str(record.get("qualified_name", "")).lower():
        score += 90
    if query_text in str(record.get("name", "")).lower():
        score += 60
    if query_text in str(record.get("signature", "")).lower():
        score += 30
    for token in query_tokens:
        if token in searchable_text:
            score += 10
    score += _path_score_adjustment(record.get("path", ""))
    return {
        "type": "symbol",
        "score": score,
        "path": record.get("path"),
        "line": record.get("line"),
        "title": record.get("qualified_name"),
        "preview": record.get("signature") or record.get("doc_summary") or record.get("kind"),
        "kind": record.get("kind"),
        "area": _file_area(record.get("path", "")),
        "changed": record.get("path") in (changed_paths or set()),
    }


def search_snapshot(snapshot, corpus, query, scope="all", limit=20, area=None, changed_only=False, path_filter=None):
    query_text = str(query or "").strip().lower()
    if not query_text:
        return []
    query_tokens = [token.lower() for token in TOKEN_PATTERN.findall(query_text)] or [query_text]
    changed_paths = {item.get("path") for item in snapshot.get("git", {}).get("changed_files", []) if item.get("path")}
    results = []

    if scope in {"all", "files"}:
        for record in snapshot.get("files", []):
            path_text = record.get("path", "")
            if not _matches_common_filters(path_text, area=area, path_filter=path_filter, changed_only=changed_only, changed_paths=changed_paths):
                continue
            file_text = corpus.get(path_text, "")
            scored = _score_text_record(record, file_text, query_tokens, query_text, changed_paths=changed_paths)
            if scored is not None:
                results.append(scored)

    if scope in {"all", "symbols"}:
        for record in snapshot.get("symbols", []):
            path_text = record.get("path", "")
            if not _matches_common_filters(path_text, area=area, path_filter=path_filter, changed_only=changed_only, changed_paths=changed_paths):
                continue
            scored = _score_symbol_record(record, query_tokens, query_text, changed_paths=changed_paths)
            if scored is not None:
                results.append(scored)

    results.sort(key=lambda item: (-item["score"], item.get("path") or "", item.get("line") or 0))
    return results[: max(1, int(limit))]


def format_search_results(results):
    if not results:
        return "No results found."
    lines = []
    for index, result in enumerate(results, start=1):
        location = result.get("path") or "(unknown)"
        if result.get("line"):
            location = f"{location}:{result['line']}"
        title = result.get("title") or "(untitled)"
        preview = result.get("preview") or ""
        kind = result.get("type")
        if result.get("kind"):
            kind = f"{kind}/{result['kind']}"
        changed_marker = " [changed]" if result.get("changed") else ""
        lines.append(f"{index}. [{kind}] {title}{changed_marker} -> {location}")
        if preview:
            lines.append(f"   {preview}")
    return "\n".join(lines)


def _filter_changed_files(snapshot, status_filter=None, area=None, path_filter=None):
    status_tokens = _normalize_status_filter(status_filter)
    filtered = []
    for item in snapshot.get("git", {}).get("changed_files", []):
        path_text = item.get("path", "")
        if area and item.get("area") != area:
            continue
        if path_filter and path_filter.lower() not in path_text.lower():
            continue
        if status_tokens and not any(token in str(item.get("status", "")).upper() for token in status_tokens):
            continue
        filtered.append(item)
    return filtered


def format_change_report(snapshot, limit=20, status_filter=None, area=None, path_filter=None, include_commits=True):
    git_snapshot = snapshot.get("git", {})
    changed_files = _filter_changed_files(snapshot, status_filter=status_filter, area=area, path_filter=path_filter)
    if not changed_files:
        return "No tracked git changes in the current snapshot for the selected filters."

    lines = [
        f"Branch: {git_snapshot.get('branch', 'unknown')}",
        f"Changed Files: {len(changed_files)} of {git_snapshot.get('changed_count', len(changed_files))}",
    ]

    status_counts = _counts_from_items(changed_files, "status")
    if status_counts:
        lines.append("Statuses: " + ", ".join(f"{status}={count}" for status, count in status_counts.items()))
    area_counts = _counts_from_items(changed_files, "area")
    if area_counts:
        lines.append("Areas: " + ", ".join(f"{area_name}={count}" for area_name, count in area_counts.items()))

    for item in changed_files[: max(1, int(limit))]:
        source_path = item.get("source_path")
        if source_path:
            lines.append(f"- {item.get('status', '??')} [{item.get('area', 'root')}]: {source_path} -> {item.get('path', '')}")
        else:
            lines.append(f"- {item.get('status', '??')} [{item.get('area', 'root')}]: {item.get('path', '')}")
    remaining = len(changed_files) - min(len(changed_files), max(1, int(limit)))
    if remaining > 0:
        lines.append(f"- ... {remaining} more")

    if include_commits:
        commits = git_snapshot.get("recent_commits", [])
        if commits:
            lines.append("Recent Commits:")
            for commit in commits[:3]:
                lines.append(
                    f"- {commit.get('short_commit', '')} {commit.get('date', '')} {commit.get('subject', '')} ({commit.get('author', '')})"
                )
    return "\n".join(lines)


def format_history_report(history_records, limit=10):
    if not history_records:
        return "No recorded librarian history yet. Run refresh first."

    lines = []
    for index, record in enumerate(reversed(history_records[-max(1, int(limit)):]), start=1):
        generated_at = str(record.get("generated_at", ""))
        branch = record.get("branch", "unknown")
        changed_count = record.get("changed_count", 0)
        area_counts = record.get("area_counts", {})
        areas_text = ", ".join(f"{area_name}={count}" for area_name, count in list(area_counts.items())[:3]) or "none"
        lines.append(f"{index}. {generated_at} | branch={branch} | changed={changed_count} | areas={areas_text}")
    return "\n".join(lines)


def format_workspace_stats(workspace):
    summary = workspace.snapshot.get("summary", {})
    git_snapshot = workspace.snapshot.get("git", {})
    lines = [
        f"Repo Root: {workspace.repo_root}",
        f"Files: {summary.get('files', 0)}",
        f"Symbols: {summary.get('symbols', 0)}",
        f"Bytes Indexed: {summary.get('bytes', 0)}",
        f"Branch: {git_snapshot.get('branch', 'unknown')}",
        f"Changed Files: {summary.get('changed_files', 0)}",
        f"History Entries: {summary.get('history_entries', len(workspace.history))}",
    ]
    if git_snapshot.get("status_counts"):
        lines.append(
            "Status Counts: "
            + ", ".join(f"{status}={count}" for status, count in git_snapshot.get("status_counts", {}).items())
        )
    if git_snapshot.get("area_counts"):
        lines.append(
            "Area Counts: "
            + ", ".join(f"{area_name}={count}" for area_name, count in git_snapshot.get("area_counts", {}).items())
        )
    return "\n".join(lines)


def _format_repl_welcome(workspace):
    summary = workspace.snapshot.get("summary", {})
    branch_name = workspace.snapshot.get("git", {}).get("branch", "unknown")
    changed_files = summary.get("changed_files", 0)
    lines = [
        "Project Librarian loaded in memory.",
        f"Repo: {workspace.repo_root.name} | Branch: {branch_name} | Files: {summary.get('files', 0)} | Symbols: {summary.get('symbols', 0)} | Changed: {changed_files}",
        "",
        "Examples:",
        "- search layout manager",
        "- symbols LayoutManagerController",
        "- changes",
        "- history",
        "- show README.md",
        "- docs-draft",
        "- changelog-draft",
        "- ai-models",
        "- ai-doctor",
        "- refresh",
        "- stats",
        "- help",
        "- quit",
        "",
        "Tip: use explicit CLI subcommands outside the REPL when you need filters like --changed-only, --area, or --status.",
        "",
        "Commands: search <query> | symbols <query> | files <query> | changes | history | show <path> | docs-draft | changelog-draft | ai-models | ai-doctor | refresh | stats | quit",
    ]
    return "\n".join(lines)


def _resolve_workspace_path(workspace, path_text):
    normalized = str(path_text or "").strip().replace("\\", "/")
    if not normalized:
        raise ProjectLibrarianError("A file path is required.")

    all_paths = sorted(workspace.corpus.keys())
    if normalized in workspace.corpus:
        return normalized

    suffix_matches = [candidate for candidate in all_paths if candidate.endswith(normalized)]
    if len(suffix_matches) == 1:
        return suffix_matches[0]
    if len(suffix_matches) > 1:
        raise ProjectLibrarianError(
            "Path is ambiguous. Matches: " + ", ".join(suffix_matches[:10])
        )

    contains_matches = [candidate for candidate in all_paths if normalized.lower() in candidate.lower()]
    if len(contains_matches) == 1:
        return contains_matches[0]
    if len(contains_matches) > 1:
        raise ProjectLibrarianError(
            "Path is ambiguous. Matches: " + ", ".join(contains_matches[:10])
        )
    raise ProjectLibrarianError(f"No indexed file matched '{normalized}'.")


def show_file_excerpt(workspace, path_text, query=None, line=None, context=3):
    resolved_path = _resolve_workspace_path(workspace, path_text)
    text = workspace.corpus.get(resolved_path, "")
    all_lines = text.splitlines()
    if not all_lines:
        return f"{resolved_path}\n(empty file)"

    if query:
        preview = _preview_for_query(text, query, limit=500)
        target_line = preview.get("line") or 1
    else:
        target_line = max(1, int(line or 1))

    start_line = max(1, target_line - max(0, int(context)))
    end_line = min(len(all_lines), target_line + max(0, int(context)))
    excerpt_lines = [f"{resolved_path}"]
    for line_number in range(start_line, end_line + 1):
        marker = ">" if line_number == target_line else " "
        excerpt_lines.append(f"{marker} {line_number:>4}: {all_lines[line_number - 1]}")
    return "\n".join(excerpt_lines)


def _changed_file_records(workspace, changed_only=True):
    if changed_only:
        return [workspace.file_lookup[path_text] for path_text in sorted(workspace.changed_paths) if path_text in workspace.file_lookup]
    return list(workspace.file_records)


def _records_grouped_by_area(records):
    grouped = {}
    for record in records:
        area = record.get("area") or _file_area(record.get("path", ""))
        grouped.setdefault(area, []).append(record)
    return grouped


def _summarize_record_names(records, limit=4):
    labels = []
    for record in records[:limit]:
        labels.append(record.get("path") or record.get("title") or "(unknown)")
    summary = ", ".join(labels)
    remaining = len(records) - min(len(records), limit)
    if remaining > 0:
        summary += f", and {remaining} more"
    return summary


def _draft_bullet_for_area(area_name, records, symbol_records=None):
    summary = _summarize_record_names(records)
    if area_name == "controllers":
        message = f"Updated controller workflows in {summary}."
    elif area_name == "models":
        message = f"Updated model and data-handling logic in {summary}."
    elif area_name == "views":
        message = f"Updated view and interaction behavior in {summary}."
    elif area_name == "app":
        message = f"Updated app-level services and module entry points in {summary}."
    elif area_name == "docs":
        message = f"Refreshed documentation sources in {summary}."
    elif area_name == "scripts":
        message = f"Updated project automation scripts in {summary}."
    else:
        message = f"Updated project-level files in {summary}."

    symbol_summary = _summarize_symbol_labels(symbol_records or [])
    if symbol_summary:
        return f"{message[:-1]} Key symbols touched: {symbol_summary}."
    return message


def _suggest_docs(records):
    suggestions = []
    seen = set()
    for area_name in _records_grouped_by_area(records):
        for candidate in DOC_SUGGESTIONS_BY_AREA.get(area_name, []):
            if candidate not in seen:
                suggestions.append(candidate)
                seen.add(candidate)
    return suggestions


def _ensure_output_subdir(output_dir, directory_name):
    target_dir = output_dir / directory_name
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def _write_generated_output(output_dir, directory_name, prefix, content, output_path=None, output_base_dir=None):
    if output_path:
        target_path = Path(output_path)
        if not target_path.is_absolute():
            base_dir = Path(output_base_dir or output_dir)
            target_path = (base_dir / target_path).resolve()
        target_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        target_dir = _ensure_output_subdir(output_dir, directory_name)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        target_path = target_dir / f"{prefix}_{stamp}.md"
    target_path.write_text(content, encoding="utf-8")
    return target_path


def _resolve_repo_path(repo_root, target_path):
    candidate = Path(target_path)
    if not candidate.is_absolute():
        candidate = (repo_root / candidate).resolve()
    return candidate


def _upsert_markdown_block(existing_text, content, start_marker, end_marker):
    managed_block = f"{start_marker}\n{content.rstrip()}\n{end_marker}"
    if start_marker in existing_text and end_marker in existing_text:
        pattern = re.compile(rf"{re.escape(start_marker)}.*?{re.escape(end_marker)}", re.DOTALL)
        updated_text = pattern.sub(managed_block, existing_text, count=1)
        return updated_text.rstrip() + "\n"

    if not existing_text.strip():
        return managed_block + "\n"
    return existing_text.rstrip() + "\n\n" + managed_block + "\n"


def apply_docs_update(repo_root, content, target_path=README_TARGET_NAME):
    resolved_target = _resolve_repo_path(repo_root, target_path)
    existing_text = resolved_target.read_text(encoding="utf-8") if resolved_target.exists() else ""
    updated_text = _upsert_markdown_block(existing_text, content, DOC_BLOCK_START, DOC_BLOCK_END)
    resolved_target.write_text(updated_text, encoding="utf-8")
    return resolved_target


def _upsert_changelog_entry(existing_text, entry_text, version_label):
    headings = list(CHANGELOG_HEADING_PATTERN.finditer(existing_text))
    normalized_entry = entry_text.rstrip() + "\n\n"
    for index, match in enumerate(headings):
        if match.group("version") != str(version_label):
            continue
        start_index = match.start()
        end_index = headings[index + 1].start() if index + 1 < len(headings) else len(existing_text)
        updated_text = existing_text[:start_index] + normalized_entry + existing_text[end_index:].lstrip("\n")
        return updated_text.rstrip() + "\n"

    if headings:
        insertion_index = headings[0].start()
        prefix = existing_text[:insertion_index].rstrip() + "\n\n"
        suffix = existing_text[insertion_index:].lstrip("\n")
        return (prefix + normalized_entry + suffix).rstrip() + "\n"

    if not existing_text.strip():
        return normalized_entry.rstrip() + "\n"
    return existing_text.rstrip() + "\n\n" + normalized_entry.rstrip() + "\n"


def apply_changelog_update(repo_root, content, version_label, target_path=CHANGELOG_TARGET_NAME):
    resolved_target = _resolve_repo_path(repo_root, target_path)
    existing_text = resolved_target.read_text(encoding="utf-8") if resolved_target.exists() else ""
    updated_text = _upsert_changelog_entry(existing_text, content, version_label)
    resolved_target.write_text(updated_text, encoding="utf-8")
    return resolved_target


def _parse_ollama_models(raw_output):
    model_names = []
    for line in str(raw_output or "").splitlines()[1:]:
        stripped = line.strip()
        if not stripped:
            continue
        model_name = stripped.split()[0]
        if model_name and model_name not in model_names:
            model_names.append(model_name)
    return model_names


def _select_recommended_model(models, preferred_model=None):
    preferred = str(preferred_model or "").strip()
    if preferred and preferred in models:
        return preferred

    ranked_groups = [
        [model for model in models if "qwen" in model.lower() and "coder" in model.lower()],
        [model for model in models if "qwen" in model.lower()],
        list(models),
    ]
    for group in ranked_groups:
        if group:
            return group[0]
    return None


def collect_ai_runtime_status(repo_root, preferred_model=DEFAULT_AI_MODEL, ollama_host=None):
    status = {
        "repo_root": str(repo_root),
        "ollama_host": str(ollama_host or os.environ.get("OLLAMA_HOST") or "default"),
        "preferred_model": preferred_model,
        "ollama_path": shutil.which("ollama"),
        "delegate_script": str(repo_root / "scripts" / "qwen_delegate.sh"),
        "smoke_script": str(repo_root / "scripts" / "local_ai_smoke_test.sh"),
        "models": [],
        "ollama_reachable": False,
        "preferred_model_available": False,
        "recommended_model": None,
        "issues": [],
    }

    delegate_path = Path(status["delegate_script"])
    smoke_path = Path(status["smoke_script"])
    status["delegate_script_exists"] = delegate_path.exists()
    status["delegate_script_executable"] = os.access(delegate_path, os.X_OK) if delegate_path.exists() else False
    status["smoke_script_exists"] = smoke_path.exists()

    if not status["ollama_path"]:
        status["issues"].append("ollama is not installed or is not on PATH.")
        return status

    env = os.environ.copy()
    if ollama_host:
        env["OLLAMA_HOST"] = str(ollama_host)

    completed = _run_shell_command(["ollama", "list"], cwd=repo_root, env=env)
    if completed.returncode != 0:
        error_text = (completed.stderr or completed.stdout or "Unable to query ollama.").strip()
        status["issues"].append(error_text)
        return status

    status["ollama_reachable"] = True
    status["models"] = _parse_ollama_models(completed.stdout)
    status["preferred_model_available"] = preferred_model in status["models"]
    status["recommended_model"] = _select_recommended_model(status["models"], preferred_model=preferred_model)
    if not status["models"]:
        status["issues"].append("Ollama is reachable but no local models were listed.")
    if not status["delegate_script_exists"]:
        status["issues"].append(f"AI delegate script is missing: {delegate_path}")
    elif not status["delegate_script_executable"]:
        status["issues"].append(f"AI delegate script is not executable: {delegate_path}")
    if preferred_model and not status["preferred_model_available"]:
        status["issues"].append(f"Preferred model is not available locally: {preferred_model}")
    return status


def format_ai_model_list(ai_status):
    models = ai_status.get("models", [])
    if not models:
        return "No local Ollama models were found for the selected host."
    lines = [f"Ollama Host: {ai_status.get('ollama_host', 'default')}", "Models:"]
    recommended_model = ai_status.get("recommended_model")
    preferred_model = ai_status.get("preferred_model")
    for model_name in models:
        suffix_bits = []
        if model_name == preferred_model:
            suffix_bits.append("preferred")
        if model_name == recommended_model:
            suffix_bits.append("recommended")
        suffix = f" [{' | '.join(suffix_bits)}]" if suffix_bits else ""
        lines.append(f"- {model_name}{suffix}")
    return "\n".join(lines)


def format_ai_status_report(ai_status):
    lines = [
        f"Repo Root: {ai_status.get('repo_root')}",
        f"Ollama Host: {ai_status.get('ollama_host')}",
        f"Ollama Installed: {'yes' if ai_status.get('ollama_path') else 'no'}",
        f"Ollama Reachable: {'yes' if ai_status.get('ollama_reachable') else 'no'}",
        f"Delegate Script: {ai_status.get('delegate_script')}",
        f"Delegate Executable: {'yes' if ai_status.get('delegate_script_executable') else 'no'}",
        f"Smoke Helper Present: {'yes' if ai_status.get('smoke_script_exists') else 'no'}",
        f"Preferred Model: {ai_status.get('preferred_model')}",
        f"Recommended Model: {ai_status.get('recommended_model') or '(none)'}",
    ]
    models = ai_status.get("models", [])
    if models:
        lines.append("Available Models: " + ", ".join(models))
    else:
        lines.append("Available Models: none")
    if ai_status.get("issues"):
        lines.append("Issues:")
        for issue_text in ai_status.get("issues", []):
            lines.append(f"- {issue_text}")
    else:
        lines.append("Issues: none")
    return "\n".join(lines)


def _resolve_ai_model(ai_status, requested_model):
    requested = str(requested_model or "").strip()
    if requested and requested.lower() != "auto":
        if ai_status.get("models") and requested not in ai_status.get("models", []):
            recommended_model = ai_status.get("recommended_model") or "(none)"
            available_models = ", ".join(ai_status.get("models", [])) or "none"
            raise ProjectLibrarianError(
                f"Requested Ollama model is not available locally: {requested}. Available models: {available_models}. Recommended: {recommended_model}."
            )
        return requested

    recommended_model = ai_status.get("recommended_model")
    if recommended_model:
        return recommended_model
    raise ProjectLibrarianError("No local Ollama model is available. Run 'project_librarian.py ai-models' or 'project_librarian.py ai-doctor' first.")


def _html_escape(value):
    return html.escape(str(value), quote=True)


def _parse_bool(value):
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _resolve_http_auth_token(auth_token=None):
    candidate = auth_token if auth_token is not None else os.environ.get(HTTP_AUTH_ENV_NAME)
    normalized = str(candidate or "").strip()
    if not normalized:
        return None, False
    if normalized.lower() == "auto":
        return secrets.token_urlsafe(24), True
    return normalized, False


def _request_token_candidates(request):
    authorization_header = str(request.headers.get("authorization", "")).strip()
    if authorization_header.lower().startswith("bearer "):
        yield authorization_header.split(" ", 1)[1].strip()

    header_token = str(request.headers.get(HTTP_AUTH_HEADER_NAME, "")).strip()
    if header_token:
        yield header_token

    query_token = str(request.query_params.get("token", "")).strip()
    if query_token:
        yield query_token

    cookie_token = str(request.cookies.get(HTTP_AUTH_COOKIE_NAME, "")).strip()
    if cookie_token:
        yield cookie_token


def _request_has_valid_token(request, auth_token):
    if not auth_token:
        return True
    for candidate in _request_token_candidates(request):
        if candidate and secrets.compare_digest(candidate, auth_token):
            return True
    return False


def _query_token_matches(request, auth_token):
    if not auth_token:
        return False
    query_token = str(request.query_params.get("token", "")).strip()
    return bool(query_token) and secrets.compare_digest(query_token, auth_token)


def _coerce_int(value, default, minimum=1, maximum=500):
    try:
        normalized = int(value)
    except (TypeError, ValueError):
        normalized = int(default)
    return max(minimum, min(maximum, normalized))


def _json_text_for_html(value):
    return json.dumps(value, ensure_ascii=True).replace("</", "<\\/")


def _dashboard_bootstrap_payload(service):
    return {
        "status": service.status_payload(),
        "changes": service.changes_payload(limit=12, include_commits=False),
        "history": service.history_payload(limit=5),
        "areas": list(AREA_ORDER),
        "scopes": ["all", "symbols", "text", "paths"],
        "docs_target": README_TARGET_NAME,
        "changelog_target": CHANGELOG_TARGET_NAME,
    }


def _render_login_html():
    return """<!doctype html>
<html lang='en'>
<head>
    <meta charset='utf-8'>
    <meta name='viewport' content='width=device-width, initial-scale=1'>
    <title>Project Librarian Login</title>
    <style>
        body {
            margin: 0;
            min-height: 100vh;
            display: grid;
            place-items: center;
            background: radial-gradient(circle at top, #173041 0%, #0b141b 65%);
            color: #e7f2f8;
            font-family: 'Segoe UI', sans-serif;
        }
        .panel {
            width: min(520px, calc(100vw - 32px));
            background: rgba(14, 25, 34, 0.96);
            border: 1px solid #294556;
            border-radius: 18px;
            padding: 24px;
            box-sizing: border-box;
        }
        h1 { margin: 0 0 10px; }
        p { color: #a9c0cd; line-height: 1.55; }
        label { display: block; margin: 18px 0 8px; color: #cde0e8; }
        input {
            width: 100%;
            box-sizing: border-box;
            background: #091117;
            color: #eef8fb;
            border: 1px solid #355466;
            border-radius: 12px;
            padding: 12px 14px;
        }
        button {
            margin-top: 16px;
            background: #14b8a6;
            color: #062127;
            border: none;
            border-radius: 12px;
            padding: 12px 16px;
            font-weight: 700;
            cursor: pointer;
        }
        code {
            display: inline-block;
            background: #091117;
            border: 1px solid #24404f;
            border-radius: 8px;
            padding: 2px 6px;
        }
    </style>
</head>
<body>
    <section class='panel'>
        <h1>Project Librarian</h1>
        <p>This shared server requires a token before the dashboard or MCP endpoint can be used.</p>
        <form method='get' action='/'>
            <label for='token'>Browser token</label>
            <input id='token' name='token' type='password' autocomplete='current-password' autofocus>
            <button type='submit'>Open Dashboard</button>
        </form>
        <p>MCP clients can authenticate with <code>Authorization: Bearer &lt;token&gt;</code>, the <code>""" + HTTP_AUTH_HEADER_NAME + """</code> header, or a <code>?token=...</code> query parameter.</p>
    </section>
</body>
</html>
"""


def _render_dashboard_html(auth_enabled=False, bootstrap_payload=None):
    auth_chip = "<span class='chip locked'>Token Protected</span>" if auth_enabled else "<span class='chip'>Local Only</span>"
    logout_markup = "<form method='post' action='/logout'><button type='submit'>Logout</button></form>" if auth_enabled else ""
    template = """<!doctype html>
<html lang='en'>
<head>
    <meta charset='utf-8'>
    <meta name='viewport' content='width=device-width, initial-scale=1'>
    <title>Project Librarian</title>
    <style>
        :root {
            color-scheme: dark;
            --page-bg: #091117;
            --page-grad: radial-gradient(circle at top left, #123446 0%, #091117 62%);
            --panel-bg: rgba(15, 28, 37, 0.94);
            --panel-strong: #102431;
            --panel-border: #274557;
            --text-main: #eef8fb;
            --text-muted: #94adba;
            --accent: #19c2af;
            --accent-strong: #86fff0;
            --warn: #f59e0b;
        }
        * { box-sizing: border-box; }
        body {
            margin: 0;
            min-height: 100vh;
            background: var(--page-grad);
            color: var(--text-main);
            font-family: 'Segoe UI', sans-serif;
        }
        a { color: var(--accent-strong); text-decoration: none; }
        a:hover { text-decoration: underline; }
        button, input, select {
            font: inherit;
        }
        button {
            cursor: pointer;
            border: 1px solid #315264;
            border-radius: 12px;
            background: #10202a;
            color: var(--text-main);
            padding: 10px 14px;
        }
        button.primary {
            background: var(--accent);
            color: #052128;
            border: none;
            font-weight: 700;
        }
        button.ghost {
            background: transparent;
        }
        input, select, textarea {
            width: 100%;
            border: 1px solid #355466;
            border-radius: 12px;
            background: #0b161d;
            color: var(--text-main);
            padding: 11px 12px;
        }
        textarea {
            min-height: 220px;
            resize: vertical;
        }
        .shell {
            max-width: 1440px;
            margin: 0 auto;
            padding: 24px;
        }
        .hero {
            display: flex;
            justify-content: space-between;
            gap: 18px;
            align-items: flex-start;
            margin-bottom: 22px;
        }
        .hero h1 {
            margin: 0 0 8px;
            font-size: clamp(2rem, 3vw, 2.6rem);
        }
        .hero p {
            margin: 0;
            color: var(--text-muted);
            max-width: 820px;
            line-height: 1.55;
        }
        .hero-actions {
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
            justify-content: flex-end;
            align-items: center;
        }
        .hero-actions form {
            margin: 0;
        }
        .chip {
            display: inline-flex;
            align-items: center;
            border: 1px solid #305061;
            border-radius: 999px;
            padding: 7px 12px;
            color: var(--accent-strong);
            background: rgba(18, 44, 53, 0.72);
        }
        .chip.locked {
            color: #fed7aa;
            border-color: #6b4e22;
            background: rgba(88, 57, 13, 0.45);
        }
        .statusbar {
            margin: 0 0 18px;
            padding: 12px 16px;
            border-radius: 14px;
            border: 1px solid var(--panel-border);
            background: rgba(12, 24, 32, 0.92);
            color: var(--text-muted);
        }
        .statusbar strong {
            color: var(--text-main);
        }
        .grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
            gap: 12px;
            margin-bottom: 20px;
        }
        .card, .panel {
            background: var(--panel-bg);
            border: 1px solid var(--panel-border);
            border-radius: 18px;
            box-shadow: 0 16px 48px rgba(0, 0, 0, 0.22);
        }
        .card {
            padding: 14px;
        }
        .label {
            font-size: 0.8rem;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 6px;
        }
        .value {
            font-size: 1.08rem;
            font-weight: 700;
            word-break: break-word;
        }
        .layout {
            display: grid;
            grid-template-columns: minmax(0, 1.3fr) minmax(320px, 0.9fr);
            gap: 18px;
        }
        .panel {
            padding: 18px;
            margin-bottom: 18px;
        }
        .panel h2 {
            margin: 0 0 14px;
            font-size: 1.05rem;
        }
        .controls {
            display: grid;
            grid-template-columns: minmax(0, 2fr) repeat(3, minmax(110px, 1fr));
            gap: 10px;
            align-items: end;
        }
        .controls .wide {
            grid-column: span 2;
        }
        .inline {
            display: flex;
            gap: 10px;
            align-items: center;
            flex-wrap: wrap;
        }
        .checkbox {
            display: inline-flex;
            gap: 8px;
            align-items: center;
            color: var(--text-muted);
        }
        .checkbox input {
            width: auto;
        }
        .split {
            display: grid;
            grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
            gap: 14px;
        }
        .results-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.94rem;
        }
        .results-table th,
        .results-table td {
            text-align: left;
            padding: 10px;
            border-top: 1px solid #223845;
            vertical-align: top;
        }
        .results-table thead th {
            border-top: none;
            color: var(--text-muted);
            font-weight: 600;
        }
        .results-table td button,
        .list button {
            padding: 0;
            border: none;
            background: none;
            color: var(--accent-strong);
            text-align: left;
        }
        .list {
            list-style: none;
            padding: 0;
            margin: 0;
        }
        .list li {
            padding: 10px 0;
            border-top: 1px solid #213745;
        }
        .list li:first-child {
            padding-top: 0;
            border-top: none;
        }
        .meta {
            display: block;
            margin-top: 4px;
            color: var(--text-muted);
            font-size: 0.88rem;
        }
        .preview {
            margin: 0;
            background: #071016;
            border: 1px solid #223745;
            border-radius: 14px;
            padding: 14px;
            min-height: 260px;
            overflow: auto;
            white-space: pre-wrap;
            line-height: 1.48;
        }
        .muted { color: var(--text-muted); }
        .message {
            min-height: 1.4em;
            color: var(--accent-strong);
            margin-top: 10px;
        }
        .message.error {
            color: #fca5a5;
        }
        .draft-actions,
        .panel-actions {
            display: flex;
            gap: 10px;
            align-items: center;
            flex-wrap: wrap;
            margin-top: 12px;
        }
        .footer-note {
            margin-top: 8px;
            color: var(--text-muted);
            font-size: 0.9rem;
        }
        @media (max-width: 1120px) {
            .layout,
            .split {
                grid-template-columns: 1fr;
            }
        }
        @media (max-width: 860px) {
            .hero {
                flex-direction: column;
            }
            .hero-actions {
                justify-content: flex-start;
            }
            .controls {
                grid-template-columns: 1fr;
            }
            .controls .wide {
                grid-column: span 1;
            }
        }
        .tab-nav {
            display: flex;
            gap: 2px;
            margin-bottom: 0;
            border-bottom: 2px solid #1e3647;
        }
        .tab-btn {
            background: transparent;
            border: 1px solid transparent;
            border-bottom: none;
            border-radius: 10px 10px 0 0;
            color: var(--text-muted);
            cursor: pointer;
            font-size: 0.92rem;
            font-weight: 600;
            margin-bottom: -2px;
            padding: 10px 22px;
            transition: background 0.15s, color 0.15s;
        }
        .tab-btn:hover { background: #112030; color: var(--text-main); }
        .tab-btn.active {
            background: #0f1e2a;
            border-color: #1e3647;
            border-bottom: 2px solid #0f1e2a;
            color: var(--accent-strong);
        }
        .tab-pane { display: none; padding-top: 18px; }
        .tab-pane.active { display: block; }
        .split-wide {
            display: grid;
            grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
            gap: 14px;
        }
        @media (max-width: 860px) {
            .split-wide { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>
    <div class='shell'>
        <section class='hero'>
            <div>
                <div class='inline'>
                    <h1>Project Librarian</h1>
                    __AUTH_CHIP__
                </div>
                <p>Shared browser dashboard and MCP server for the same RAM-loaded workspace. Search, file inspection, refresh, and docs or changelog drafting all use the one cached librarian service instead of spinning up a second repo load.</p>
            </div>
            <div class='hero-actions'>
                <a href='/mcp'>MCP Endpoint</a>
                <a href='/api/status'>Status JSON</a>
                <button class='primary' id='refresh-button' type='button'>Refresh Cache</button>
                __LOGOUT_BUTTON__
            </div>
        </section>

        <section class='statusbar' id='status-banner'><strong>Loading shared workspace status...</strong></section>

        <section class='grid' id='summary-cards'></section>

        <nav class='tab-nav' role='tablist'>
            <button class='tab-btn active' data-tab='librarian' role='tab' type='button'>Librarian</button>
            <button class='tab-btn' data-tab='draft-center' role='tab' type='button'>Draft Center</button>
            <button class='tab-btn' data-tab='excel-library' role='tab' type='button'>Excel Library</button>
            <button class='tab-btn' data-tab='git-controls' role='tab' type='button'>Git Controls</button>
        </nav>

        <div class='tab-pane active' id='tab-librarian'>
        <div class='layout'>
            <div>
                <section class='panel'>
                    <h2>Search</h2>
                    <form id='search-form'>
                        <div class='controls'>
                            <div>
                                <label class='label' for='search-query'>Query</label>
                                <input id='search-query' name='q' type='text' placeholder='Search symbols, paths, or indexed text'>
                            </div>
                            <div>
                                <label class='label' for='search-scope'>Scope</label>
                                <select id='search-scope' name='scope'></select>
                            </div>
                            <div>
                                <label class='label' for='search-area'>Area</label>
                                <select id='search-area' name='area'></select>
                            </div>
                            <div>
                                <label class='label' for='search-path-filter'>Path Filter</label>
                                <input id='search-path-filter' name='path_filter' type='text' placeholder='Optional path filter'>
                            </div>
                        </div>
                        <div class='panel-actions'>
                            <label class='checkbox'><input id='search-changed-only' name='changed_only' type='checkbox'>Changed only</label>
                            <button class='primary' type='submit'>Run Search</button>
                            <button class='ghost' id='search-clear' type='button'>Clear</button>
                        </div>
                    </form>
                    <div class='message' id='search-message'></div>
                </section>

                <section class='panel'>
                    <h2>Search Results</h2>
                    <div id='search-results'><p class='muted'>Run a search to inspect indexed text and symbols without reloading the repository.</p></div>
                </section>

                <section class='panel'>
                    <div class='inline' style='justify-content: space-between;'>
                        <h2>File Excerpt</h2>
                        <span class='muted' id='excerpt-path'></span>
                    </div>
                    <pre class='preview' id='file-excerpt'>Select a changed file or search result to inspect an excerpt here.</pre>
                </section>
            </div>

            <div>
                <section class='panel'>
                    <h2>Changed Files</h2>
                    <ul class='list' id='changes-list'></ul>
                </section>

                <section class='panel'>
                    <h2>Recent Refreshes</h2>
                    <ul class='list' id='history-list'></ul>
                </section>
            </div>
        </div>
        </div>

        <div class='tab-pane' id='tab-draft-center'>
        <section class='panel'>
            <h2>Draft Center</h2>
            <div class='split-wide'>
                <div>
                    <div class='inline' style='justify-content: space-between;'>
                        <h3 style='margin: 0;'>Documentation Draft</h3>
                        <label class='checkbox'><input id='docs-changed-only' type='checkbox' checked>Changed only</label>
                    </div>
                    <div class='controls' style='grid-template-columns: 1fr 1fr;'>
                        <div>
                            <label class='label' for='docs-title'>Title</label>
                            <input id='docs-title' type='text' placeholder='Project Documentation Update Draft'>
                        </div>
                        <div>
                            <label class='label' for='docs-target'>Apply Target</label>
                            <input id='docs-target' type='text'>
                        </div>
                    </div>
                    <div class='draft-actions'>
                        <button class='primary' id='docs-preview' type='button'>Preview Draft</button>
                        <button id='docs-save' type='button'>Write Draft File</button>
                        <button id='docs-apply' type='button'>Apply to Target</button>
                    </div>
                    <div class='message' id='docs-message'></div>
                    <pre class='preview' id='docs-preview-text'>No documentation draft generated yet.</pre>
                </div>
                <div>
                    <div class='inline' style='justify-content: space-between;'>
                        <h3 style='margin: 0;'>Changelog Draft</h3>
                        <label class='checkbox'><input id='changelog-changed-only' type='checkbox' checked>Changed only</label>
                    </div>
                    <div class='controls' style='grid-template-columns: 1fr 1fr 1fr;'>
                        <div>
                            <label class='label' for='changelog-version'>Version</label>
                            <input id='changelog-version' type='text' placeholder='Unreleased'>
                        </div>
                        <div>
                            <label class='label' for='changelog-date'>Date</label>
                            <input id='changelog-date' type='text' placeholder='YYYY-MM-DD'>
                        </div>
                        <div>
                            <label class='label' for='changelog-target'>Apply Target</label>
                            <input id='changelog-target' type='text'>
                        </div>
                    </div>
                    <div class='draft-actions'>
                        <button class='primary' id='changelog-preview' type='button'>Preview Draft</button>
                        <button id='changelog-save' type='button'>Write Draft File</button>
                        <button id='changelog-apply' type='button'>Apply to Target</button>
                    </div>
                    <div class='message' id='changelog-message'></div>
                    <pre class='preview' id='changelog-preview-text'>No changelog draft generated yet.</pre>
                </div>
            </div>
            <p class='footer-note'>Draft previews are generated from the shared cached workspace and can either be written to draft artifacts or applied back to README and CHANGELOG managed sections.</p>
        </section>
        </div>

        <div class='tab-pane' id='tab-excel-library'>
        <section class='panel'>
            <h2>Excel Library</h2>
            <p class='muted' style='margin: 0 0 14px;'>Configure a folder to browse Excel and CSV files. Choose keyword columns (header names like Shop Order/Part Number/Date or column letters like A,B,D) to power search indexing.</p>
            <div class='controls' style='grid-template-columns: 1fr 2fr auto; align-items: end;'>
                <div>
                    <label class='label' for='excel-library-name'>Library Name</label>
                    <input id='excel-library-name' type='text' placeholder='My Excel Library'>
                </div>
                <div>
                    <label class='label' for='excel-folder-path'>Folder Path</label>
                    <input id='excel-folder-path' type='text' placeholder='/absolute/path/to/excel/folder'>
                </div>
                <button class='primary' id='excel-save-config' type='button'>Save &amp; Scan</button>
            </div>
            <div class='controls' style='grid-template-columns: 1fr; margin-top: 10px;'>
                <div>
                    <label class='label' for='excel-keyword-columns'>Keyword Columns</label>
                    <input id='excel-keyword-columns' type='text' placeholder='Downtime Code, Shop Order, Part Number, Date'>
                </div>
            </div>
            <div class='message' id='excel-message'></div>
        </section>

        <section class='panel' id='excel-files-panel' style='display: none;'>
            <div class='inline' style='justify-content: space-between; margin-bottom: 14px;'>
                <h2 id='excel-panel-title' style='margin: 0;'>Excel Files</h2>
                <button class='ghost' id='excel-rescan' type='button'>Rescan</button>
            </div>
            <div id='excel-files-list'></div>
        </section>

        <section class='panel'>
            <div class='inline' style='justify-content: space-between; margin-bottom: 12px;'>
                <h2 style='margin: 0;'>Excel Keyword Search</h2>
            </div>
            <div class='controls' style='grid-template-columns: minmax(0, 2fr) auto auto; align-items: end;'>
                <div>
                    <label class='label' for='excel-search-query'>Query</label>
                    <input id='excel-search-query' type='text' placeholder='Search downtime code, shop order, part number, or date'>
                </div>
                <button class='primary' id='excel-search-run' type='button'>Search</button>
                <button class='ghost' id='excel-search-clear' type='button'>Clear</button>
            </div>
            <div class='message' id='excel-search-message'></div>
            <div id='excel-search-results'><p class='muted'>Search results will appear here.</p></div>
        </section>
        </div>

        <div class='tab-pane' id='tab-git-controls'>
        <section class='panel'>
            <div class='inline' style='justify-content: space-between; margin-bottom: 10px;'>
                <h2 style='margin: 0;'>Working Tree</h2>
                <div class='inline'>
                    <button class='ghost' id='git-refresh' type='button'>Refresh</button>
                    <button class='ghost' id='git-stage-all' type='button'>Stage All</button>
                    <button class='ghost' id='git-unstage-all' type='button'>Unstage All</button>
                </div>
            </div>
            <div id='git-status-summary' class='muted'>Git status not loaded yet.</div>
            <div id='git-changes-list' style='margin-top: 12px;'></div>
            <div class='message' id='git-message'></div>
        </section>

        <section class='panel'>
            <div class='split-wide'>
                <div>
                    <h2 style='margin-top: 0;'>Commit &amp; Push</h2>
                    <div class='controls' style='grid-template-columns: 1fr;'>
                        <div>
                            <label class='label' for='git-commit-message'>Commit Message</label>
                            <input id='git-commit-message' type='text' placeholder='Describe what changed'>
                        </div>
                    </div>
                    <div class='draft-actions'>
                        <button class='primary' id='git-commit-run' type='button'>Commit</button>
                        <button id='git-review-head' type='button'>Review HEAD</button>
                    </div>

                    <div class='controls' style='grid-template-columns: 1fr 1fr auto; margin-top: 12px;'>
                        <div>
                            <label class='label' for='git-remote'>Remote</label>
                            <input id='git-remote' type='text' value='origin'>
                        </div>
                        <div>
                            <label class='label' for='git-branch'>Branch</label>
                            <input id='git-branch' type='text' placeholder='Current branch'>
                        </div>
                        <button class='primary' id='git-push-run' type='button'>Push</button>
                    </div>

                    <h3 style='margin-bottom: 10px;'>Recent Commits</h3>
                    <ul class='list' id='git-commits-list'></ul>
                </div>
                <div>
                    <div class='inline' style='justify-content: space-between;'>
                        <h2 style='margin: 0;'>Review</h2>
                        <span id='git-review-label' class='muted'></span>
                    </div>
                    <pre class='preview' id='git-review-text'>Select a file diff or commit review target.</pre>
                </div>
            </div>
        </section>
        </div>

    </div>

    <script id='initial-state' type='application/json'>__INITIAL_STATE__</script>
    <script>
        const state = {
            bootstrap: JSON.parse(document.getElementById('initial-state').textContent),
            lastSearch: null,
            lastPath: '',
            lastQueryForExcerpt: '',
        };

        const byId = (id) => document.getElementById(id);
        const cardsEl = byId('summary-cards');
        const changesEl = byId('changes-list');
        const historyEl = byId('history-list');
        const searchResultsEl = byId('search-results');
        const excerptEl = byId('file-excerpt');
        const excerptPathEl = byId('excerpt-path');
        const statusBannerEl = byId('status-banner');
        const searchMessageEl = byId('search-message');
        const docsMessageEl = byId('docs-message');
        const changelogMessageEl = byId('changelog-message');
        const excelMessageEl = byId('excel-message');
        const excelSearchMessageEl = byId('excel-search-message');
        const gitMessageEl = byId('git-message');

        function setMessage(element, text, isError = false) {
            element.textContent = text || '';
            element.classList.toggle('error', Boolean(isError));
        }

        function escapeHtml(value) {
            return String(value ?? '')
                .replaceAll('&', '&amp;')
                .replaceAll('<', '&lt;')
                .replaceAll('>', '&gt;')
                .replaceAll('"', '&quot;')
                .replaceAll("'", '&#39;');
        }

        async function fetchJson(url, options = {}) {
            const headers = { Accept: 'application/json', ...(options.headers || {}) };
            if (options.body && !headers['Content-Type']) {
                headers['Content-Type'] = 'application/json';
            }
            const response = await fetch(url, { credentials: 'same-origin', ...options, headers });
            const contentType = response.headers.get('content-type') || '';
            const payload = contentType.includes('application/json') ? await response.json() : await response.text();
            if (!response.ok) {
                const message = typeof payload === 'string' ? payload : payload.error || payload.message || `Request failed with status ${response.status}`;
                throw new Error(message);
            }
            return payload;
        }

        function renderSummary(status) {
            const cards = [
                ['Branch', status.branch || 'unknown'],
                ['Files', status.files || 0],
                ['Symbols', status.symbols || 0],
                ['Changed Files', status.changed_files || 0],
                ['History Entries', status.history_entries || 0],
                ['Last Refresh', status.last_refresh_at || 'unknown'],
            ];
            if (status.last_refresh_error) {
                cards.push(['Refresh Error', status.last_refresh_error]);
            }
            cardsEl.innerHTML = cards.map(([label, value]) => `
                <div class="card">
                    <div class="label">${escapeHtml(label)}</div>
                    <div class="value">${escapeHtml(value)}</div>
                </div>
            `).join('');

            const workerText = status.refresh_worker_running ? 'running' : 'manual';
            const watcherText = status.library_watcher?.watcher_running ? 'active' : 'inactive';
            const watchedFiles = status.library_watcher?.tracked_files || 0;
            statusBannerEl.innerHTML = `<strong>${escapeHtml(status.repo_root || '')}</strong> on branch <strong>${escapeHtml(status.branch || 'unknown')}</strong>. Refresh worker: <strong>${escapeHtml(workerText)}</strong> at ${escapeHtml(status.refresh_interval_seconds || 0)}s. Library auto-watcher: <strong>${escapeHtml(watcherText)}</strong> · ${escapeHtml(watchedFiles)} tracked.`;
        }

        function renderChanges(payload) {
            const items = payload.changes || [];
            if (!items.length) {
                changesEl.innerHTML = '<li class="muted">No tracked changes.</li>';
                return;
            }
            changesEl.innerHTML = items.map((item) => `
                <li>
                    <button type="button" class="open-file" data-path="${escapeHtml(item.path || '')}">${escapeHtml(item.path || '')}</button>
                    <span class="meta">${escapeHtml(item.status || '?')} · ${escapeHtml(item.area || 'unknown')}</span>
                </li>
            `).join('');
        }

        function renderHistory(payload) {
            const items = payload.history || [];
            if (!items.length) {
                historyEl.innerHTML = '<li class="muted">No refresh history yet.</li>';
                return;
            }
            historyEl.innerHTML = items.map((item) => `
                <li>
                    <strong>${escapeHtml(item.generated_at || 'unknown')}</strong>
                    <span class="meta">${escapeHtml(item.summary?.changed_files || 0)} changed files · ${escapeHtml(item.summary?.symbols || 0)} symbols</span>
                </li>
            `).join('');
        }

        function renderSearchResults(payload) {
            const items = payload.results || [];
            if (!items.length) {
                searchResultsEl.innerHTML = '<p class="muted">No results for the current query.</p>';
                return;
            }
            searchResultsEl.innerHTML = `
                <table class="results-table">
                    <thead>
                        <tr><th>Path</th><th>Scope</th><th>Area</th><th>Line</th><th>Preview</th></tr>
                    </thead>
                    <tbody>
                        ${items.map((item) => `
                            <tr>
                                <td><button type="button" class="open-file" data-path="${escapeHtml(item.path || '')}">${escapeHtml(item.path || '')}</button></td>
                                <td>${escapeHtml(item.scope || '')}</td>
                                <td>${escapeHtml(item.area || '')}</td>
                                <td>${escapeHtml(item.line || '')}</td>
                                <td>${escapeHtml(item.preview || '')}</td>
                            </tr>
                        `).join('')}
                    </tbody>
                </table>
            `;
        }

        function renderDraftPreview(targetId, payload, actionLabel) {
            byId(targetId).textContent = payload.content || '';
            const location = payload.applied_path || payload.output_path || '';
            return location ? `${actionLabel}: ${location}` : actionLabel;
        }

        async function loadOverview() {
            const [status, changes, history] = await Promise.all([
                fetchJson('/api/status'),
                fetchJson('/api/changes?limit=12&include_commits=0'),
                fetchJson('/api/history?limit=5'),
            ]);
            state.bootstrap.status = status;
            state.bootstrap.changes = changes;
            state.bootstrap.history = history;
            renderSummary(status);
            renderChanges(changes);
            renderHistory(history);
        }

        async function runSearch() {
            const query = byId('search-query').value.trim();
            const scope = byId('search-scope').value;
            const area = byId('search-area').value;
            const pathFilter = byId('search-path-filter').value.trim();
            const changedOnly = byId('search-changed-only').checked;
            if (!query) {
                searchResultsEl.innerHTML = '<p class="muted">Run a search to inspect indexed text and symbols without reloading the repository.</p>';
                setMessage(searchMessageEl, '');
                state.lastSearch = null;
                return;
            }

            const params = new URLSearchParams({ q: query, scope });
            if (area) {
                params.set('area', area);
            }
            if (pathFilter) {
                params.set('path_filter', pathFilter);
            }
            if (changedOnly) {
                params.set('changed_only', '1');
            }
            const payload = await fetchJson(`/api/search?${params.toString()}`);
            state.lastSearch = { query, scope, area, pathFilter, changedOnly };
            renderSearchResults(payload);
            setMessage(searchMessageEl, `${payload.count || 0} result(s) loaded.`);
        }

        async function loadExcerpt(pathValue, query = '') {
            if (!pathValue) {
                return;
            }
            const params = new URLSearchParams({ path: pathValue });
            if (query) {
                params.set('query', query);
            }
            const payload = await fetchJson(`/api/file?${params.toString()}`);
            state.lastPath = pathValue;
            state.lastQueryForExcerpt = query;
            excerptPathEl.textContent = pathValue;
            excerptEl.textContent = payload.excerpt || '';
        }

        async function refreshWorkspace() {
            await fetchJson('/api/refresh', { method: 'POST', body: JSON.stringify({}) });
            await loadOverview();
            if (state.lastSearch) {
                await runSearch();
            }
            if (state.lastPath) {
                await loadExcerpt(state.lastPath, state.lastQueryForExcerpt);
            }
        }

        async function handleDraft(endpoint, targetId, messageEl, payload, successText) {
            const response = await fetchJson(endpoint, {
                method: 'POST',
                body: JSON.stringify(payload),
            });
            setMessage(messageEl, renderDraftPreview(targetId, response, successText));
            return response;
        }

        function _formatBytes(bytes) {
            if (bytes < 1024) return `${bytes} B`;
            if (bytes < 1048576) return `${(bytes / 1024).toFixed(1)} KB`;
            return `${(bytes / 1048576).toFixed(1)} MB`;
        }

        function renderExcelFiles(payload) {
            const panel = byId('excel-files-panel');
            const listEl = byId('excel-files-list');
            const titleEl = byId('excel-panel-title');
            if (payload.error && !(payload.files && payload.files.length)) {
                setMessage(excelMessageEl, payload.error, true);
                panel.style.display = 'none';
                return;
            }
            const files = payload.files || [];
            const folder = payload.folder || '';
            setMessage(excelMessageEl, `${files.length} file(s) found in ${folder || '(no folder selected)'}.`);
            titleEl.textContent = `Excel Files (${files.length})`;
            panel.style.display = '';
            if (!files.length) {
                listEl.innerHTML = '<p class="muted">No Excel or CSV files found in this folder.</p>';
                return;
            }
            listEl.innerHTML = `
                <table class="results-table">
                    <thead>
                        <tr><th>File</th><th>Sheets</th><th>Size</th><th>Modified (UTC)</th></tr>
                    </thead>
                    <tbody>
                        ${files.map((f) => `
                            <tr>
                                <td>${escapeHtml(f.name || '')}</td>
                                <td>${f.sheets && f.sheets.length ? escapeHtml(f.sheets.join(', ')) : '<span class="muted">—</span>'}</td>
                                <td>${escapeHtml(_formatBytes(f.size_bytes || 0))}</td>
                                <td>${escapeHtml((f.modified_at || '').replace('T', ' ').replace('Z', ''))}</td>
                            </tr>
                        `).join('')}
                    </tbody>
                </table>
            `;
        }

        function renderExcelSearchResults(payload) {
            const target = byId('excel-search-results');
            const items = payload.results || [];
            if (!items.length) {
                target.innerHTML = '<p class="muted">No matches found for the current query.</p>';
                return;
            }
            target.innerHTML = `
                <table class="results-table">
                    <thead>
                        <tr><th>File</th><th>Sheet</th><th>Row</th><th>Field</th><th>Value</th></tr>
                    </thead>
                    <tbody>
                        ${items.map((item) => `
                            <tr>
                                <td>${escapeHtml(item.file || '')}</td>
                                <td>${escapeHtml(item.sheet || '')}</td>
                                <td>${escapeHtml(item.row || '')}</td>
                                <td>${escapeHtml(item.field || '')}</td>
                                <td>${escapeHtml(item.value || '')}</td>
                            </tr>
                        `).join('')}
                    </tbody>
                </table>
            `;
        }

        function _gitHasStagedChanges(change) {
            const token = String(change.xy || '  ');
            return token.length >= 1 && token[0] !== ' ' && token[0] !== '?';
        }

        function _gitHasUnstagedChanges(change) {
            const token = String(change.xy || '  ');
            return (token.length >= 2 && token[1] !== ' ') || token.includes('?');
        }

        function renderGitStatus(payload) {
            const summaryEl = byId('git-status-summary');
            const changesEl = byId('git-changes-list');
            const commitsEl = byId('git-commits-list');
            const branch = payload.branch || 'unknown';
            const upstream = payload.upstream || '(no upstream)';
            byId('git-branch').value = byId('git-branch').value.trim() || branch;

            summaryEl.innerHTML = `
                <strong>Branch:</strong> ${escapeHtml(branch)}
                · <strong>Upstream:</strong> ${escapeHtml(upstream)}
                · <strong>Ahead:</strong> ${escapeHtml(payload.ahead || 0)}
                · <strong>Behind:</strong> ${escapeHtml(payload.behind || 0)}
                · <strong>Changed:</strong> ${escapeHtml(payload.changed_count || 0)}
            `;

            const changes = payload.changed_files || [];
            if (!changes.length) {
                changesEl.innerHTML = '<p class="muted">No local changes in working tree.</p>';
            } else {
                changesEl.innerHTML = `
                    <table class="results-table">
                        <thead>
                            <tr><th>Status</th><th>Path</th><th>Actions</th></tr>
                        </thead>
                        <tbody>
                            ${changes.map((item) => {
                                const canStage = _gitHasUnstagedChanges(item);
                                const canUnstage = _gitHasStagedChanges(item);
                                const path = escapeHtml(item.path || '');
                                const status = escapeHtml(item.status || item.xy || '??');
                                return `
                                    <tr>
                                        <td>${status}</td>
                                        <td>${path}</td>
                                        <td>
                                            ${canStage ? `<button type="button" class="git-stage-file" data-path="${path}">stage</button>` : '<span class="muted">—</span>'}
                                            ${canUnstage ? `<button type="button" class="git-unstage-file" data-path="${path}">unstage</button>` : ''}
                                            <button type="button" class="git-diff-file" data-path="${path}">diff</button>
                                        </td>
                                    </tr>
                                `;
                            }).join('')}
                        </tbody>
                    </table>
                `;
            }

            const commits = payload.recent_commits || [];
            if (!commits.length) {
                commitsEl.innerHTML = '<li class="muted">No commits available.</li>';
                return;
            }
            commitsEl.innerHTML = commits.map((commit) => `
                <li>
                    <button type="button" class="git-review-commit" data-commit="${escapeHtml(commit.commit || '')}">
                        ${escapeHtml(commit.short_commit || '')} · ${escapeHtml(commit.subject || '')}
                    </button>
                    <span class="meta">${escapeHtml(commit.author || '')} · ${escapeHtml(commit.date || '')}</span>
                </li>
            `).join('');
        }

        function renderGitReview(label, textContent) {
            byId('git-review-label').textContent = label || '';
            byId('git-review-text').textContent = textContent || '';
        }

        async function loadGitStatus() {
            const payload = await fetchJson('/api/git/status?limit=20');
            renderGitStatus(payload);
            return payload;
        }

        async function loadGitDiff(pathValue) {
            const params = new URLSearchParams();
            if (pathValue) {
                params.set('path', pathValue);
            }
            const payload = await fetchJson(`/api/git/diff?${params.toString()}`);
            renderGitReview(`Diff: ${pathValue || 'working tree'}`, payload.diff || '(no diff)');
        }

        async function loadGitCommitReview(commitValue = 'HEAD') {
            const params = new URLSearchParams({ commit: commitValue });
            const payload = await fetchJson(`/api/git/commit-review?${params.toString()}`);
            renderGitReview(`Commit: ${payload.commit || commitValue}`, payload.review || '(empty commit review)');
        }

        async function runGitAction(endpoint, payload, successText) {
            const response = await fetchJson(endpoint, {
                method: 'POST',
                body: JSON.stringify(payload || {}),
            });
            setMessage(gitMessageEl, successText || response.message || 'Git command complete.');
            await loadGitStatus();
            return response;
        }

        async function loadExcelFiles() {
            const payload = await fetchJson('/api/excel-files');
            renderExcelFiles(payload);
        }

        async function runExcelSearch() {
            const query = byId('excel-search-query').value.trim();
            if (!query) {
                byId('excel-search-results').innerHTML = '<p class="muted">Search results will appear here.</p>';
                setMessage(excelSearchMessageEl, '');
                return;
            }
            const params = new URLSearchParams({ q: query, limit: '200' });
            const payload = await fetchJson(`/api/excel-search?${params.toString()}`);
            renderExcelSearchResults(payload);
            setMessage(excelSearchMessageEl, `${payload.count || 0} match(es) for "${query}".`);
        }

        async function saveExcelConfig() {
            const name = byId('excel-library-name').value.trim();
            const folder = byId('excel-folder-path').value.trim();
            const keywordColumns = byId('excel-keyword-columns').value.trim();
            if (!folder) {
                setMessage(excelMessageEl, 'Please enter a folder path.', true);
                return;
            }
            const config = await fetchJson('/api/excel-config', {
                method: 'POST',
                body: JSON.stringify({
                    name: name || 'Excel Library',
                    folder,
                    keyword_columns: keywordColumns,
                }),
            });
            setMessage(excelMessageEl, `Saved. Scanning "${config.folder}"…`);
            await loadExcelFiles();
        }

        async function loadExcelLibrary() {
            try {
                const config = await fetchJson('/api/excel-config');
                byId('excel-library-name').value = config.name || '';
                byId('excel-folder-path').value = config.folder || '';
                const keywords = Array.isArray(config.keyword_columns) ? config.keyword_columns.join(', ') : '';
                byId('excel-keyword-columns').value = keywords;
                if (config.folder) {
                    await loadExcelFiles();
                }
            } catch (error) {
                setMessage(excelMessageEl, error.message, true);
            }
        }

        function installSelectOptions() {
            const scopeSelect = byId('search-scope');
            scopeSelect.innerHTML = state.bootstrap.scopes.map((scope) => `<option value="${escapeHtml(scope)}">${escapeHtml(scope)}</option>`).join('');

            const areaSelect = byId('search-area');
            areaSelect.innerHTML = [''].concat(state.bootstrap.areas || []).map((area) => {
                const label = area || 'all';
                return `<option value="${escapeHtml(area)}">${escapeHtml(label)}</option>`;
            }).join('');

            byId('docs-target').value = state.bootstrap.docs_target || 'README.md';
            byId('changelog-target').value = state.bootstrap.changelog_target || 'CHANGELOG.md';
        }

        function installEvents() {
            byId('search-form').addEventListener('submit', async (event) => {
                event.preventDefault();
                try {
                    await runSearch();
                } catch (error) {
                    setMessage(searchMessageEl, error.message, true);
                }
            });

            byId('search-clear').addEventListener('click', () => {
                byId('search-query').value = '';
                byId('search-path-filter').value = '';
                byId('search-area').value = '';
                byId('search-scope').value = 'all';
                byId('search-changed-only').checked = false;
                searchResultsEl.innerHTML = '<p class="muted">Run a search to inspect indexed text and symbols without reloading the repository.</p>';
                setMessage(searchMessageEl, '');
                state.lastSearch = null;
            });

            document.addEventListener('click', async (event) => {
                const button = event.target.closest('.open-file');
                if (!button) {
                    return;
                }
                try {
                    await loadExcerpt(button.dataset.path || '', byId('search-query').value.trim());
                } catch (error) {
                    excerptEl.textContent = error.message;
                    excerptPathEl.textContent = button.dataset.path || '';
                }
            });

            document.addEventListener('click', async (event) => {
                const stageBtn = event.target.closest('.git-stage-file');
                const unstageBtn = event.target.closest('.git-unstage-file');
                const diffBtn = event.target.closest('.git-diff-file');
                const reviewBtn = event.target.closest('.git-review-commit');
                if (!stageBtn && !unstageBtn && !diffBtn && !reviewBtn) {
                    return;
                }
                try {
                    if (stageBtn) {
                        await runGitAction('/api/git/stage', { path: stageBtn.dataset.path || '' }, `Staged ${stageBtn.dataset.path || ''}`);
                    } else if (unstageBtn) {
                        await runGitAction('/api/git/unstage', { path: unstageBtn.dataset.path || '' }, `Unstaged ${unstageBtn.dataset.path || ''}`);
                    } else if (diffBtn) {
                        await loadGitDiff(diffBtn.dataset.path || '');
                    } else if (reviewBtn) {
                        await loadGitCommitReview(reviewBtn.dataset.commit || 'HEAD');
                    }
                } catch (error) {
                    setMessage(gitMessageEl, error.message, true);
                }
            });

            byId('refresh-button').addEventListener('click', async () => {
                try {
                    setMessage(searchMessageEl, 'Refreshing shared cache...');
                    await refreshWorkspace();
                    setMessage(searchMessageEl, 'Shared cache refreshed.');
                } catch (error) {
                    setMessage(searchMessageEl, error.message, true);
                }
            });

            byId('docs-preview').addEventListener('click', async () => {
                try {
                    await handleDraft('/api/docs-draft', 'docs-preview-text', docsMessageEl, {
                        title: byId('docs-title').value.trim() || null,
                        changed_only: byId('docs-changed-only').checked,
                        apply: false,
                        target_path: byId('docs-target').value.trim() || 'README.md',
                    }, 'Documentation draft generated');
                } catch (error) {
                    setMessage(docsMessageEl, error.message, true);
                }
            });

            byId('docs-save').addEventListener('click', async () => {
                try {
                    await handleDraft('/api/docs-draft', 'docs-preview-text', docsMessageEl, {
                        title: byId('docs-title').value.trim() || null,
                        changed_only: byId('docs-changed-only').checked,
                        apply: false,
                        target_path: byId('docs-target').value.trim() || 'README.md',
                    }, 'Documentation draft written');
                } catch (error) {
                    setMessage(docsMessageEl, error.message, true);
                }
            });

            byId('docs-apply').addEventListener('click', async () => {
                try {
                    await handleDraft('/api/docs-draft', 'docs-preview-text', docsMessageEl, {
                        title: byId('docs-title').value.trim() || null,
                        changed_only: byId('docs-changed-only').checked,
                        apply: true,
                        target_path: byId('docs-target').value.trim() || 'README.md',
                    }, 'Documentation draft applied');
                } catch (error) {
                    setMessage(docsMessageEl, error.message, true);
                }
            });

            byId('changelog-preview').addEventListener('click', async () => {
                try {
                    await handleDraft('/api/changelog-draft', 'changelog-preview-text', changelogMessageEl, {
                        version_text: byId('changelog-version').value.trim() || null,
                        release_date: byId('changelog-date').value.trim() || null,
                        changed_only: byId('changelog-changed-only').checked,
                        apply: false,
                        target_path: byId('changelog-target').value.trim() || 'CHANGELOG.md',
                    }, 'Changelog draft generated');
                } catch (error) {
                    setMessage(changelogMessageEl, error.message, true);
                }
            });

            byId('changelog-save').addEventListener('click', async () => {
                try {
                    await handleDraft('/api/changelog-draft', 'changelog-preview-text', changelogMessageEl, {
                        version_text: byId('changelog-version').value.trim() || null,
                        release_date: byId('changelog-date').value.trim() || null,
                        changed_only: byId('changelog-changed-only').checked,
                        apply: false,
                        target_path: byId('changelog-target').value.trim() || 'CHANGELOG.md',
                    }, 'Changelog draft written');
                } catch (error) {
                    setMessage(changelogMessageEl, error.message, true);
                }
            });

            byId('changelog-apply').addEventListener('click', async () => {
                try {
                    await handleDraft('/api/changelog-draft', 'changelog-preview-text', changelogMessageEl, {
                        version_text: byId('changelog-version').value.trim() || null,
                        release_date: byId('changelog-date').value.trim() || null,
                        changed_only: byId('changelog-changed-only').checked,
                        apply: true,
                        target_path: byId('changelog-target').value.trim() || 'CHANGELOG.md',
                    }, 'Changelog draft applied');
                } catch (error) {
                    setMessage(changelogMessageEl, error.message, true);
                }
            });
            document.querySelectorAll('.tab-btn').forEach((btn) => {
                btn.addEventListener('click', () => {
                    document.querySelectorAll('.tab-btn').forEach((b) => b.classList.remove('active'));
                    document.querySelectorAll('.tab-pane').forEach((p) => p.classList.remove('active'));
                    btn.classList.add('active');
                    const pane = byId(`tab-${btn.dataset.tab}`);
                    if (pane) pane.classList.add('active');
                    if (btn.dataset.tab === 'excel-library') {
                        loadExcelLibrary().catch((error) => setMessage(excelMessageEl, error.message, true));
                    }
                    if (btn.dataset.tab === 'git-controls') {
                        loadGitStatus().catch((error) => setMessage(gitMessageEl, error.message, true));
                    }
                });
            });

            byId('excel-save-config').addEventListener('click', async () => {
                try {
                    await saveExcelConfig();
                } catch (error) {
                    setMessage(excelMessageEl, error.message, true);
                }
            });

            byId('excel-rescan').addEventListener('click', async () => {
                try {
                    setMessage(excelMessageEl, 'Scanning…');
                    await loadExcelFiles();
                } catch (error) {
                    setMessage(excelMessageEl, error.message, true);
                }
            });

            byId('excel-search-run').addEventListener('click', async () => {
                try {
                    await runExcelSearch();
                } catch (error) {
                    setMessage(excelSearchMessageEl, error.message, true);
                }
            });

            byId('excel-search-clear').addEventListener('click', () => {
                byId('excel-search-query').value = '';
                byId('excel-search-results').innerHTML = '<p class="muted">Search results will appear here.</p>';
                setMessage(excelSearchMessageEl, '');
            });

            byId('excel-search-query').addEventListener('keydown', async (event) => {
                if (event.key !== 'Enter') {
                    return;
                }
                event.preventDefault();
                try {
                    await runExcelSearch();
                } catch (error) {
                    setMessage(excelSearchMessageEl, error.message, true);
                }
            });

            byId('git-refresh').addEventListener('click', async () => {
                try {
                    await loadGitStatus();
                    setMessage(gitMessageEl, 'Git status refreshed.');
                } catch (error) {
                    setMessage(gitMessageEl, error.message, true);
                }
            });

            byId('git-stage-all').addEventListener('click', async () => {
                try {
                    await runGitAction('/api/git/stage', { all: true }, 'All changes staged.');
                } catch (error) {
                    setMessage(gitMessageEl, error.message, true);
                }
            });

            byId('git-unstage-all').addEventListener('click', async () => {
                try {
                    await runGitAction('/api/git/unstage', { all: true }, 'All staged changes moved back to working tree.');
                } catch (error) {
                    setMessage(gitMessageEl, error.message, true);
                }
            });

            byId('git-review-head').addEventListener('click', async () => {
                try {
                    await loadGitCommitReview('HEAD');
                } catch (error) {
                    setMessage(gitMessageEl, error.message, true);
                }
            });

            byId('git-commit-run').addEventListener('click', async () => {
                const message = byId('git-commit-message').value.trim();
                if (!message) {
                    setMessage(gitMessageEl, 'Commit message is required.', true);
                    return;
                }
                try {
                    const payload = await runGitAction('/api/git/commit', { message }, 'Commit created.');
                    byId('git-commit-message').value = '';
                    if (payload && payload.commit) {
                        await loadGitCommitReview(payload.commit);
                    }
                } catch (error) {
                    setMessage(gitMessageEl, error.message, true);
                }
            });

            byId('git-push-run').addEventListener('click', async () => {
                const remote = byId('git-remote').value.trim() || 'origin';
                const branch = byId('git-branch').value.trim();
                if (!branch) {
                    setMessage(gitMessageEl, 'Branch is required before push.', true);
                    return;
                }
                try {
                    await runGitAction('/api/git/push', { remote, branch }, `Pushed ${branch} to ${remote}.`);
                } catch (error) {
                    setMessage(gitMessageEl, error.message, true);
                }
            });
        }

        async function start() {
            installSelectOptions();
            renderSummary(state.bootstrap.status || {});
            renderChanges(state.bootstrap.changes || {});
            renderHistory(state.bootstrap.history || {});
            installEvents();
            if (window.location.search.includes('token=')) {
                window.history.replaceState({}, document.title, '/');
            }
        }

        start().catch((error) => {
            statusBannerEl.textContent = error.message;
            statusBannerEl.classList.add('error');
        });
    </script>
</body>
</html>
"""
    return (
        template.replace("__AUTH_CHIP__", auth_chip)
        .replace("__LOGOUT_BUTTON__", logout_markup)
        .replace("__INITIAL_STATE__", _json_text_for_html(bootstrap_payload or {}))
    )


def _load_excel_library_config(output_dir):
    config_path = Path(output_dir) / EXCEL_LIBRARY_CONFIG_NAME
    if config_path.exists():
        try:
            loaded = json.loads(config_path.read_text(encoding="utf-8"))
            if not isinstance(loaded, dict):
                loaded = {}
            loaded.setdefault("folder", "")
            loaded.setdefault("name", "Excel Library")
            loaded.setdefault("keyword_columns", list(DEFAULT_EXCEL_KEYWORD_COLUMNS))
            loaded["keyword_columns"] = _normalize_keyword_columns(loaded.get("keyword_columns")) or list(DEFAULT_EXCEL_KEYWORD_COLUMNS)
            return loaded
        except Exception:
            pass
    return {
        "folder": "",
        "name": "Excel Library",
        "keyword_columns": list(DEFAULT_EXCEL_KEYWORD_COLUMNS),
    }


def _save_excel_library_config(output_dir, config):
    config_path = Path(output_dir) / EXCEL_LIBRARY_CONFIG_NAME
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(json.dumps(config, indent=2), encoding="utf-8")


def _normalize_keyword_columns(raw_value):
    if isinstance(raw_value, str):
        pieces = [part.strip() for part in raw_value.split(",")]
    elif isinstance(raw_value, list):
        pieces = [str(part).strip() for part in raw_value]
    else:
        pieces = []
    return [piece for piece in pieces if piece]


def _column_letter_to_index(value):
    token = str(value or "").strip().upper()
    if not token or not token.isalpha():
        return None
    index = 0
    for char in token:
        index = (index * 26) + (ord(char) - 64)
    return index - 1


def _normalize_excel_value(cell_value):
    if cell_value is None:
        return ""
    if isinstance(cell_value, datetime):
        return f"{cell_value.date().isoformat()} {cell_value.strftime('%m/%d/%Y')}"
    if isinstance(cell_value, date):
        return f"{cell_value.isoformat()} {cell_value.strftime('%m/%d/%Y')}"
    text = str(cell_value).strip()
    if not text:
        return ""
    return text


def _resolve_keyword_column_indices(keyword_columns, header_row):
    if not keyword_columns:
        return []
    header_lookup = {}
    for index, value in enumerate(header_row or []):
        key = str(value or "").strip().lower()
        if key:
            header_lookup[key] = index
    indices = []
    for token in keyword_columns:
        col_index = _column_letter_to_index(token)
        if col_index is not None:
            indices.append(col_index)
            continue
        key = str(token).strip().lower()
        if key in header_lookup:
            indices.append(header_lookup[key])
    return sorted(set(indices))


def _iter_excel_keyword_rows(folder_path, keyword_columns):
    folder = Path(folder_path)
    if not folder.is_dir():
        return
    workbook_paths = sorted(
        (path for path in folder.rglob("*") if path.is_file() and path.suffix.lower() in EXCEL_EXTENSIONS),
        key=lambda p: str(p.relative_to(folder)).lower(),
    )
    for workbook_path in workbook_paths:
        suffix = workbook_path.suffix.lower()
        if suffix in (".xlsx", ".xlsm"):
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(str(workbook_path), read_only=True, data_only=True)
            except Exception:
                continue
            try:
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    rows = worksheet.iter_rows(values_only=True)
                    header_row = next(rows, None)
                    keyword_indices = _resolve_keyword_column_indices(keyword_columns, header_row)
                    for row_number, row in enumerate(rows, start=2):
                        if not row:
                            continue
                        selected_indices = keyword_indices or range(len(row))
                        fields = []
                        for idx in selected_indices:
                            if idx < 0 or idx >= len(row):
                                continue
                            cell_text = _normalize_excel_value(row[idx])
                            if not cell_text:
                                continue
                            header_name = ""
                            if header_row and idx < len(header_row):
                                header_name = str(header_row[idx] or "").strip()
                            if not header_name:
                                header_name = f"Column {idx + 1}"
                            fields.append({"field": header_name, "value": cell_text})
                        if fields:
                            yield {
                                "file": str(workbook_path),
                                "sheet": sheet_name,
                                "row": row_number,
                                "fields": fields,
                            }
            finally:
                workbook.close()
        elif suffix == ".csv":
            try:
                with workbook_path.open("r", encoding="utf-8", newline="") as handle:
                    reader = csv.reader(handle)
                    header_row = next(reader, None)
                    keyword_indices = _resolve_keyword_column_indices(keyword_columns, header_row)
                    for row_number, row in enumerate(reader, start=2):
                        if not row:
                            continue
                        selected_indices = keyword_indices or range(len(row))
                        fields = []
                        for idx in selected_indices:
                            if idx < 0 or idx >= len(row):
                                continue
                            cell_text = _normalize_excel_value(row[idx])
                            if not cell_text:
                                continue
                            header_name = ""
                            if header_row and idx < len(header_row):
                                header_name = str(header_row[idx] or "").strip()
                            if not header_name:
                                header_name = f"Column {idx + 1}"
                            fields.append({"field": header_name, "value": cell_text})
                        if fields:
                            yield {
                                "file": str(workbook_path),
                                "sheet": "CSV",
                                "row": row_number,
                                "fields": fields,
                            }
            except Exception:
                continue


def _search_excel_rows(folder_path, keyword_columns, query, limit=200):
    query_text = str(query or "").strip().lower()
    if not query_text:
        return []
    query_tokens = [token for token in re.split(r"\s+", query_text) if token]
    if not query_tokens:
        return []

    results = []
    for row_payload in _iter_excel_keyword_rows(folder_path, keyword_columns):
        for field in row_payload["fields"]:
            value_text = field["value"]
            haystack = value_text.lower()
            if all(token in haystack for token in query_tokens):
                results.append(
                    {
                        "file": row_payload["file"],
                        "sheet": row_payload["sheet"],
                        "row": row_payload["row"],
                        "field": field["field"],
                        "value": value_text,
                    }
                )
                if len(results) >= limit:
                    return results
    return results


def _list_excel_files(folder_path):
    folder = Path(folder_path)
    if not folder.is_dir():
        return []
    results = []
    for entry in sorted((path for path in folder.rglob("*") if path.is_file()), key=lambda p: str(p.relative_to(folder)).lower()):
        if entry.suffix.lower() not in EXCEL_EXTENSIONS:
            continue
        stat = entry.stat()
        file_info = {
            "name": entry.name,
            "path": str(entry),
            "size_bytes": stat.st_size,
            "modified_at": _format_iso_utc(stat.st_mtime),
            "sheets": [],
        }
        if entry.suffix.lower() in (".xlsx", ".xlsm"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(entry), read_only=True, data_only=True)
                file_info["sheets"] = list(wb.sheetnames)
                wb.close()
            except Exception:
                pass
        results.append(file_info)
    return results


def _format_iso_utc(timestamp_float):
    return datetime.fromtimestamp(timestamp_float, tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _wrap_http_app_with_token_auth(app, auth_token):
    if not auth_token:
        return app

    from starlette.requests import Request
    from starlette.responses import HTMLResponse, JSONResponse, PlainTextResponse

    class _ProtectedHttpApp:
        def __init__(self, wrapped_app, required_token):
            self._wrapped_app = wrapped_app
            self._required_token = required_token

        async def __call__(self, scope, receive, send):
            if scope.get("type") != "http":
                await self._wrapped_app(scope, receive, send)
                return

            path = str(scope.get("path") or "")
            method = str(scope.get("method") or "GET").upper()
            if (path == "/" and method == "GET") or (path == "/logout" and method == "POST"):
                await self._wrapped_app(scope, receive, send)
                return

            request = Request(scope, receive=receive)
            if _request_has_valid_token(request, self._required_token):
                await self._wrapped_app(scope, receive, send)
                return

            if path.startswith("/api/"):
                response = JSONResponse(
                    {
                        "error": (
                            "Missing or invalid Project Librarian token. Provide Authorization: Bearer <token>, "
                            f"the {HTTP_AUTH_HEADER_NAME} header, or ?token=<token>."
                        )
                    },
                    status_code=401,
                )
            elif path.startswith("/mcp") or path.startswith("/sse") or path.startswith("/messages"):
                response = PlainTextResponse(
                    (
                        "Unauthorized. Provide Authorization: Bearer <token>, "
                        f"the {HTTP_AUTH_HEADER_NAME} header, or ?token=<token>."
                    ),
                    status_code=401,
                )
            else:
                response = HTMLResponse(_render_login_html(), status_code=401)

            await response(scope, receive, send)

    return _ProtectedHttpApp(app, auth_token)


def create_librarian_mcp_server(
    service,
    host=DEFAULT_MCP_HOST,
    port=DEFAULT_MCP_PORT,
    log_level="INFO",
    auth_token=None,
    transport=DEFAULT_MCP_TRANSPORT,
):
    try:
        from mcp.server.fastmcp import FastMCP
    except ImportError as exc:
        raise ProjectLibrarianError(
            "The 'mcp' package is required for mcp-server mode. Install it in the active environment first."
        ) from exc
    from starlette.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse

    server = FastMCP(
        name="Project Librarian",
        instructions=(
            "Persistent in-memory repository librarian for search, change tracking, draft generation, and local AI diagnostics. "
            "This server keeps one cached workspace instance alive so multiple MCP clients can share it without repeated reloads."
        ),
        host=host,
        port=port,
        log_level=log_level,
    )

    async def _request_json_payload(request):
        try:
            payload = await request.json()
        except json.JSONDecodeError as exc:
            raise ProjectLibrarianError("Request body must be valid JSON.") from exc
        if not isinstance(payload, dict):
            raise ProjectLibrarianError("Request body must be a JSON object.")
        return payload

    normalized_transport = str(transport or DEFAULT_MCP_TRANSPORT).strip().lower()

    def _mcp_probe_payload(request=None):
        base_url = ""
        if request is not None:
            try:
                base_url = str(request.base_url).rstrip("/")
            except Exception:
                base_url = ""
        endpoint_map = {
            "streamable-http": f"{base_url}/mcp" if base_url else "/mcp",
            "sse": f"{base_url}/sse" if base_url else "/sse",
            "messages": f"{base_url}/messages" if base_url else "/messages",
            "probe": f"{base_url}/api/mcp-probe" if base_url else "/api/mcp-probe",
            "probe_sse": f"{base_url}/api/mcp-probe/sse" if base_url else "/api/mcp-probe/sse",
            "probe_jsonrpc": f"{base_url}/api/mcp-probe/jsonrpc" if base_url else "/api/mcp-probe/jsonrpc",
        }
        return {
            "status": "ok",
            "server": "Project Librarian",
            "auth_required": bool(auth_token),
            "transport_configured": normalized_transport,
            "endpoints": endpoint_map,
            "auth": {
                "accepted": ["Authorization: Bearer <token>", f"{HTTP_AUTH_HEADER_NAME}: <token>", "?token=<token>"],
            },
            "jsonrpc_probe_example": {
                "method": "POST",
                "url": endpoint_map["probe_jsonrpc"],
                "body": {"jsonrpc": "2.0", "id": "probe-1", "method": "mcp.ping", "params": {}},
            },
            "sse_probe_example": {
                "method": "GET",
                "url": endpoint_map["probe_sse"],
                "accept": "text/event-stream",
            },
        }

    @server.custom_route("/api/mcp-probe", methods=["GET"], include_in_schema=False)
    async def _mcp_probe_route(request):
        return JSONResponse(_mcp_probe_payload(request))

    @server.custom_route("/api/mcp-probe/sse", methods=["GET"], include_in_schema=False)
    async def _mcp_probe_sse_route(request):
        payload = _mcp_probe_payload(request)

        async def _probe_stream():
            yield "event: ready\n"
            yield f"data: {json.dumps(payload)}\n\n"
            yield "event: ping\n"
            yield "data: mcp-sse-probe\n\n"

        return StreamingResponse(_probe_stream(), media_type="text/event-stream")

    @server.custom_route("/api/mcp-probe/jsonrpc", methods=["POST"], include_in_schema=False)
    async def _mcp_probe_jsonrpc_route(request):
        try:
            payload = await _request_json_payload(request)
        except Exception as exc:
            return JSONResponse(
                {
                    "jsonrpc": "2.0",
                    "id": None,
                    "error": {"code": -32700, "message": f"Invalid JSON payload: {exc}"},
                },
                status_code=400,
            )

        request_id = payload.get("id")
        method = str(payload.get("method") or "").strip()
        if not method:
            return JSONResponse(
                {
                    "jsonrpc": "2.0",
                    "id": request_id,
                    "error": {"code": -32600, "message": "Missing JSON-RPC method."},
                },
                status_code=400,
            )

        if method in {"mcp.ping", "ping", "rpc.ping"}:
            return JSONResponse(
                {
                    "jsonrpc": "2.0",
                    "id": request_id,
                    "result": {
                        "ok": True,
                        "server": "Project Librarian",
                        "transport_configured": normalized_transport,
                        "auth_required": bool(auth_token),
                    },
                }
            )

        if method in {"mcp.probe", "rpc.discover", "probe"}:
            return JSONResponse(
                {
                    "jsonrpc": "2.0",
                    "id": request_id,
                    "result": _mcp_probe_payload(request),
                }
            )

        return JSONResponse(
            {
                "jsonrpc": "2.0",
                "id": request_id,
                "error": {"code": -32601, "message": f"Unsupported probe method: {method}"},
            },
            status_code=404,
        )

    @server.custom_route("/", methods=["GET"], include_in_schema=False)
    async def _dashboard_route(request):
        if auth_token and not _request_has_valid_token(request, auth_token):
            return HTMLResponse(_render_login_html(), status_code=401)

        response = HTMLResponse(
            _render_dashboard_html(
                auth_enabled=bool(auth_token),
                bootstrap_payload=_dashboard_bootstrap_payload(service),
            )
        )
        if auth_token and _query_token_matches(request, auth_token):
            response.set_cookie(
                HTTP_AUTH_COOKIE_NAME,
                auth_token,
                httponly=True,
                samesite="strict",
                max_age=86400,
                path="/",
            )
        return response

    @server.custom_route("/logout", methods=["POST"], include_in_schema=False)
    async def _dashboard_logout_route(request):
        response = RedirectResponse(url="/", status_code=303)
        response.delete_cookie(HTTP_AUTH_COOKIE_NAME, path="/")
        return response

    @server.custom_route("/api/status", methods=["GET"], include_in_schema=False)
    async def _dashboard_status_route(request):
        payload = service.status_payload()
        payload["auth_enabled"] = bool(auth_token)
        return JSONResponse(payload)

    @server.custom_route("/api/changes", methods=["GET"], include_in_schema=False)
    async def _dashboard_changes_route(request):
        limit = _coerce_int(request.query_params.get("limit"), 12, minimum=1, maximum=200)
        status_filter = request.query_params.get("status_filter") or None
        area = request.query_params.get("area") or None
        path_filter = request.query_params.get("path_filter") or None
        include_commits = _parse_bool(request.query_params.get("include_commits", "0"))
        return JSONResponse(
            service.changes_payload(
                limit=limit,
                status_filter=status_filter,
                area=area,
                path_filter=path_filter,
                include_commits=include_commits,
            )
        )

    @server.custom_route("/api/history", methods=["GET"], include_in_schema=False)
    async def _dashboard_history_route(request):
        limit = _coerce_int(request.query_params.get("limit"), 5, minimum=1, maximum=100)
        return JSONResponse(service.history_payload(limit=limit))

    @server.custom_route("/api/search", methods=["GET"], include_in_schema=False)
    async def _dashboard_search_route(request):
        query = str(request.query_params.get("q", "")).strip()
        if not query:
            return JSONResponse({"count": 0, "results": [], "formatted": ""})
        scope = request.query_params.get("scope", "all") or "all"
        area = request.query_params.get("area") or None
        changed_only = _parse_bool(request.query_params.get("changed_only"))
        path_filter = request.query_params.get("path_filter") or None
        limit = _coerce_int(request.query_params.get("limit"), 20, minimum=1, maximum=100)
        return JSONResponse(
            service.search_payload(
                query,
                scope=scope,
                limit=limit,
                area=area,
                changed_only=changed_only,
                path_filter=path_filter,
            )
        )

    @server.custom_route("/api/file", methods=["GET"], include_in_schema=False)
    async def _dashboard_file_route(request):
        path_text = str(request.query_params.get("path", "")).strip()
        if not path_text:
            return JSONResponse({"error": "The 'path' query parameter is required."}, status_code=400)
        query = request.query_params.get("query") or None
        line_value = request.query_params.get("line") or None
        line_number = None if not line_value else _coerce_int(line_value, line_value, minimum=1, maximum=500000)
        context = _coerce_int(request.query_params.get("context"), 5, minimum=1, maximum=40)
        try:
            excerpt = service.show_excerpt(path_text, query=query, line=line_number, context=context)
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=404)
        return JSONResponse({"path": path_text, "excerpt": excerpt})

    @server.custom_route("/api/refresh", methods=["POST"], include_in_schema=False)
    async def _dashboard_refresh_api_route(request):
        refresh_result = service.refresh()
        return JSONResponse(
            {
                "status": service.status_payload(),
                "summary": refresh_result["snapshot"].get("summary", {}),
                "snapshot_path": str(refresh_result["snapshot_path"]),
            }
        )

    @server.custom_route("/refresh", methods=["POST"], include_in_schema=False)
    async def _dashboard_refresh_route(request):
        service.refresh()
        return RedirectResponse(url="/", status_code=303)

    @server.custom_route("/api/docs-draft", methods=["POST"], include_in_schema=False)
    async def _dashboard_docs_draft_route(request):
        try:
            payload = await _request_json_payload(request)
            return JSONResponse(
                service.docs_draft_payload(
                    title=(payload.get("title") or None),
                    changed_only=_parse_bool(payload.get("changed_only", True)),
                    output_path=(payload.get("output_path") or None),
                    apply=_parse_bool(payload.get("apply", False)),
                    target_path=str(payload.get("target_path") or README_TARGET_NAME),
                )
            )
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/changelog-draft", methods=["POST"], include_in_schema=False)
    async def _dashboard_changelog_draft_route(request):
        try:
            payload = await _request_json_payload(request)
            return JSONResponse(
                service.changelog_draft_payload(
                    version_text=(payload.get("version_text") or None),
                    release_date=(payload.get("release_date") or None),
                    changed_only=_parse_bool(payload.get("changed_only", True)),
                    output_path=(payload.get("output_path") or None),
                    apply=_parse_bool(payload.get("apply", False)),
                    target_path=str(payload.get("target_path") or CHANGELOG_TARGET_NAME),
                )
            )
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/excel-config", methods=["GET"], include_in_schema=False)
    async def _excel_config_get_route(request):
        config = _load_excel_library_config(service.output_dir)
        return JSONResponse(config)

    @server.custom_route("/api/excel-config", methods=["POST"], include_in_schema=False)
    async def _excel_config_save_route(request):
        try:
            payload = await _request_json_payload(request)
            config = _load_excel_library_config(service.output_dir)
            if "folder" in payload:
                config["folder"] = str(payload["folder"]).strip()
            if "name" in payload:
                config["name"] = str(payload["name"]).strip() or "Excel Library"
            if "keyword_columns" in payload:
                config["keyword_columns"] = _normalize_keyword_columns(payload.get("keyword_columns"))
                if not config["keyword_columns"]:
                    config["keyword_columns"] = list(DEFAULT_EXCEL_KEYWORD_COLUMNS)
            _save_excel_library_config(service.output_dir, config)
            return JSONResponse(config)
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/excel-files", methods=["GET"], include_in_schema=False)
    async def _excel_files_route(request):
        config = _load_excel_library_config(service.output_dir)
        folder = config.get("folder", "").strip()
        if not folder:
            return JSONResponse({"files": [], "folder": "", "count": 0, "error": "No folder configured. Use the Excel Library tab to set a folder path."})
        try:
            files = _list_excel_files(folder)
            return JSONResponse({"files": files, "folder": folder, "count": len(files)})
        except Exception as exc:
            return JSONResponse({"files": [], "folder": folder, "count": 0, "error": str(exc)}, status_code=400)

    @server.custom_route("/api/excel-search", methods=["GET"], include_in_schema=False)
    async def _excel_search_route(request):
        config = _load_excel_library_config(service.output_dir)
        folder = str(config.get("folder") or "").strip()
        if not folder:
            return JSONResponse({"results": [], "count": 0, "error": "No folder configured. Use the Excel Library tab to set a folder path."}, status_code=400)
        query = str(request.query_params.get("q") or "").strip()
        if not query:
            return JSONResponse({"results": [], "count": 0, "query": ""})
        limit = _coerce_int(request.query_params.get("limit"), 200, minimum=1, maximum=1000)
        keyword_columns = _normalize_keyword_columns(config.get("keyword_columns"))
        try:
            results = _search_excel_rows(folder, keyword_columns, query, limit=limit)
            return JSONResponse(
                {
                    "results": results,
                    "count": len(results),
                    "query": query,
                    "folder": folder,
                    "keyword_columns": keyword_columns,
                }
            )
        except Exception as exc:
            return JSONResponse({"results": [], "count": 0, "error": str(exc)}, status_code=400)

    @server.custom_route("/api/git/status", methods=["GET"], include_in_schema=False)
    async def _git_status_route(request):
        limit = _coerce_int(request.query_params.get("limit"), 20, minimum=1, maximum=100)
        try:
            return JSONResponse(_collect_git_status_payload(service.repo_root, commit_limit=limit))
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/git/diff", methods=["GET"], include_in_schema=False)
    async def _git_diff_route(request):
        path_text = str(request.query_params.get("path") or "").strip()
        staged = _parse_bool(request.query_params.get("staged", "0"))
        try:
            return JSONResponse(
                {
                    "path": path_text,
                    "staged": staged,
                    "diff": _read_git_diff_text(service.repo_root, path_text=path_text or None, staged=staged),
                }
            )
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/git/commit-review", methods=["GET"], include_in_schema=False)
    async def _git_commit_review_route(request):
        commit_ref = str(request.query_params.get("commit") or "HEAD").strip()
        if not commit_ref:
            commit_ref = "HEAD"
        try:
            return JSONResponse(
                {
                    "commit": commit_ref,
                    "review": _read_git_commit_review(service.repo_root, commit_ref),
                }
            )
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/git/stage", methods=["POST"], include_in_schema=False)
    async def _git_stage_route(request):
        try:
            payload = await _request_json_payload(request)
            if _parse_bool(payload.get("all")):
                _run_git_action(service.repo_root, ["add", "-A"])
                message = "Staged all changes."
            else:
                path_text = str(payload.get("path") or "").strip()
                if not path_text:
                    raise ProjectLibrarianError("'path' is required when 'all' is false.")
                _run_git_action(service.repo_root, ["add", "--", path_text])
                message = f"Staged {path_text}."
            return JSONResponse({"ok": True, "message": message, "git": _collect_git_status_payload(service.repo_root, commit_limit=20)})
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/git/unstage", methods=["POST"], include_in_schema=False)
    async def _git_unstage_route(request):
        try:
            payload = await _request_json_payload(request)
            if _parse_bool(payload.get("all")):
                _run_git_action(service.repo_root, ["reset"])
                message = "Moved all staged changes back to working tree."
            else:
                path_text = str(payload.get("path") or "").strip()
                if not path_text:
                    raise ProjectLibrarianError("'path' is required when 'all' is false.")
                _run_git_action(service.repo_root, ["reset", "HEAD", "--", path_text])
                message = f"Unstaged {path_text}."
            return JSONResponse({"ok": True, "message": message, "git": _collect_git_status_payload(service.repo_root, commit_limit=20)})
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/git/commit", methods=["POST"], include_in_schema=False)
    async def _git_commit_route(request):
        try:
            payload = await _request_json_payload(request)
            message_text = str(payload.get("message") or "").strip()
            if not message_text:
                raise ProjectLibrarianError("Commit message is required.")
            _run_git_action(service.repo_root, ["commit", "-m", message_text])
            latest = _collect_recent_commits(service.repo_root, limit=1)
            commit_ref = latest[0].get("commit") if latest else ""
            return JSONResponse(
                {
                    "ok": True,
                    "message": "Commit created.",
                    "commit": commit_ref,
                    "recent_commits": _collect_recent_commits(service.repo_root, limit=20),
                    "git": _collect_git_status_payload(service.repo_root, commit_limit=20),
                }
            )
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.custom_route("/api/git/push", methods=["POST"], include_in_schema=False)
    async def _git_push_route(request):
        try:
            payload = await _request_json_payload(request)
            remote_name = str(payload.get("remote") or "origin").strip() or "origin"
            branch_name = str(payload.get("branch") or "").strip()
            if not branch_name:
                branch_name = _run_git_command(service.repo_root, "rev-parse", "--abbrev-ref", "HEAD") or ""
            if not branch_name:
                raise ProjectLibrarianError("Unable to determine branch for push.")
            _run_git_action(service.repo_root, ["push", remote_name, branch_name])
            return JSONResponse(
                {
                    "ok": True,
                    "message": f"Pushed {branch_name} to {remote_name}.",
                    "git": _collect_git_status_payload(service.repo_root, commit_limit=20),
                }
            )
        except Exception as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)

    @server.resource("librarian://status", name="librarian-status", description="Current persistent librarian service status.")
    def _status_resource() -> dict[str, object]:
        return service.status_payload()

    @server.resource("librarian://stats", name="librarian-stats", description="Human-readable librarian workspace statistics.")
    def _stats_resource() -> str:
        return service.stats_text()

    @server.resource("librarian://snapshot", name="librarian-snapshot", description="Current librarian snapshot payload.")
    def _snapshot_resource() -> dict[str, object]:
        return service.snapshot_payload()

    @server.tool(name="refresh", description="Refresh the shared in-memory librarian workspace and on-disk snapshot.", structured_output=True)
    def _refresh_tool() -> dict[str, object]:
        result = service.refresh()
        summary = result["snapshot"]["summary"]
        return {
            "summary": summary,
            "snapshot_path": str(result["snapshot_path"]),
            "history_path": str(result["history_path"]),
            "corpus_path": str(result["corpus_path"]),
        }

    @server.tool(name="status", description="Return the shared service status for the persistent librarian instance.", structured_output=True)
    def _status_tool() -> dict[str, object]:
        return service.status_payload()

    @server.tool(name="stats", description="Return human-readable librarian workspace statistics.")
    def _stats_tool() -> str:
        return service.stats_text()

    @server.tool(name="search", description="Search the cached librarian workspace without reloading it.", structured_output=True)
    def _search_tool(
        query: str,
        scope: str = "all",
        limit: int = 20,
        area: str | None = None,
        changed_only: bool = False,
        path_filter: str | None = None,
    ) -> dict[str, object]:
        return service.search_payload(
            query,
            scope=scope,
            limit=limit,
            area=area,
            changed_only=changed_only,
            path_filter=path_filter,
        )

    @server.tool(name="changes", description="Report tracked git changes from the cached librarian snapshot.", structured_output=True)
    def _changes_tool(
        limit: int = 20,
        status_filter: str | None = None,
        area: str | None = None,
        path_filter: str | None = None,
        include_commits: bool = True,
    ) -> dict[str, object]:
        return service.changes_payload(
            limit=limit,
            status_filter=status_filter,
            area=area,
            path_filter=path_filter,
            include_commits=include_commits,
        )

    @server.tool(name="history", description="Return recent librarian refresh history from the shared workspace.", structured_output=True)
    def _history_tool(limit: int = 10) -> dict[str, object]:
        return service.history_payload(limit=limit)

    @server.tool(name="show_file", description="Show a file excerpt from the cached RAM-loaded corpus.")
    def _show_tool(path_text: str, query: str | None = None, line: int | None = None, context: int = 3) -> str:
        return service.show_excerpt(path_text, query=query, line=line, context=context)

    @server.tool(name="docs_draft", description="Generate or apply a documentation draft using the shared workspace.", structured_output=True)
    def _docs_draft_tool(
        title: str | None = None,
        changed_only: bool = True,
        output_path: str | None = None,
        apply: bool = False,
        target_path: str = README_TARGET_NAME,
    ) -> dict[str, object]:
        return service.docs_draft_payload(
            title=title,
            changed_only=changed_only,
            output_path=output_path,
            apply=apply,
            target_path=target_path,
        )

    @server.tool(name="changelog_draft", description="Generate or apply a changelog draft using the shared workspace.", structured_output=True)
    def _changelog_draft_tool(
        version_text: str | None = None,
        release_date: str | None = None,
        changed_only: bool = True,
        output_path: str | None = None,
        apply: bool = False,
        target_path: str = CHANGELOG_TARGET_NAME,
    ) -> dict[str, object]:
        return service.changelog_draft_payload(
            version_text=version_text,
            release_date=release_date,
            changed_only=changed_only,
            output_path=output_path,
            apply=apply,
            target_path=target_path,
        )

    @server.tool(name="ai_models", description="List Ollama models visible to the persistent librarian service.", structured_output=True)
    def _ai_models_tool(preferred_model: str = DEFAULT_AI_MODEL, ollama_host: str | None = None) -> dict[str, object]:
        return service.ai_models_payload(preferred_model=preferred_model, ollama_host=ollama_host)

    @server.tool(name="ai_doctor", description="Diagnose local Ollama and delegate readiness for the persistent librarian service.", structured_output=True)
    def _ai_doctor_tool(preferred_model: str = DEFAULT_AI_MODEL, ollama_host: str | None = None) -> dict[str, object]:
        return service.ai_doctor_payload(preferred_model=preferred_model, ollama_host=ollama_host)

    @server.tool(name="ai_summary", description="Run the optional local AI summary against the shared librarian workspace.", structured_output=True)
    def _ai_summary_tool(
        task: str = "Summarize the current repository changes and likely next actions.",
        mode: str = "analysis",
        model: str = DEFAULT_AI_MODEL,
        changed_only: bool = True,
        ollama_host: str | None = None,
    ) -> dict[str, object]:
        return service.ai_summary_payload(
            task=task,
            mode=mode,
            model=model,
            changed_only=changed_only,
            ollama_host=ollama_host,
        )

    return server


def run_librarian_mcp_server(
    repo_root=None,
    output_dir=None,
    transport=DEFAULT_MCP_TRANSPORT,
    host=DEFAULT_MCP_HOST,
    port=DEFAULT_MCP_PORT,
    refresh_first=False,
    refresh_interval_seconds=DEFAULT_REFRESH_INTERVAL_SECONDS,
    log_level="INFO",
    auth_token=None,
):
    resolved_auth_token, generated_auth_token = _resolve_http_auth_token(auth_token)
    service = LibrarianService(
        repo_root=repo_root,
        output_dir=output_dir,
        refresh_if_missing=True,
        refresh_first=refresh_first,
        refresh_interval_seconds=refresh_interval_seconds,
        start_refresh_worker=refresh_interval_seconds > 0,
    )
    server = create_librarian_mcp_server(
        service,
        host=host,
        port=port,
        log_level=log_level,
        auth_token=resolved_auth_token,
        transport=transport,
    )

    if transport == "streamable-http":
        print(f"Project Librarian dashboard at http://{host}:{port}/")
        print(f"Project Librarian MCP server listening at http://{host}:{port}/mcp")
    elif transport == "sse":
        print(f"Project Librarian MCP server listening at http://{host}:{port}/sse")
    else:
        print("Project Librarian MCP server running on stdio transport")
    if resolved_auth_token and transport != "stdio":
        browser_hint = f"http://{host}:{port}/?token={resolved_auth_token}"
        print(f"HTTP auth token enabled via {HTTP_AUTH_ENV_NAME if auth_token is None else '--auth-token'}")
        print(f"Dashboard login URL: {browser_hint}")
        print(f"MCP auth header: Authorization: Bearer {resolved_auth_token}")
        if generated_auth_token:
            print("Token mode: generated for this process")
    if transport in {"streamable-http", "sse"}:
        print(f"MCP probe JSON: http://{host}:{port}/api/mcp-probe")
        print(f"MCP probe SSE: http://{host}:{port}/api/mcp-probe/sse")
        print(f"MCP probe JSON-RPC: http://{host}:{port}/api/mcp-probe/jsonrpc")
    if refresh_interval_seconds > 0:
        print(f"Background refresh worker interval: {refresh_interval_seconds:.1f}s")
    else:
        print("Background refresh worker disabled")

    try:
        if transport == "stdio":
            server.run(transport=transport)
        else:
            try:
                import uvicorn
            except ImportError as exc:
                raise ProjectLibrarianError(
                    "The 'uvicorn' package is required for HTTP MCP server mode. Install it in the active environment first."
                ) from exc

            if transport == "streamable-http":
                app = server.streamable_http_app()
            elif transport == "sse":
                app = server.sse_app()
            else:
                raise ProjectLibrarianError(f"Unsupported MCP transport: {transport}")
            app = _wrap_http_app_with_token_auth(app, resolved_auth_token)
            uvicorn.run(app, host=host, port=port, log_level=log_level.lower())
    finally:
        service.stop()


def generate_docs_draft(workspace, title=None, changed_only=True):
    records = _changed_file_records(workspace, changed_only=changed_only)
    grouped = _records_grouped_by_area(records)
    touched_symbols = _collect_touched_symbols(workspace, records)
    area_symbols = _collect_area_symbol_summary(records, touched_symbols)
    commit_subjects = _recent_commit_subjects(workspace)
    branch_name = workspace.snapshot.get("git", {}).get("branch", "unknown")
    lines = [
        f"# {title or 'Project Documentation Update Draft'}",
        "",
        f"- Generated: {workspace.snapshot.get('generated_at', _utc_now_text())}",
        f"- Branch: {branch_name}",
        f"- Scope: {'changed files only' if changed_only else 'full indexed workspace'}",
        "",
        "## Summary",
        "",
        f"- Indexed files: {workspace.snapshot.get('summary', {}).get('files', 0)}",
        f"- Indexed symbols: {workspace.snapshot.get('summary', {}).get('symbols', 0)}",
        f"- Changed files in snapshot: {workspace.snapshot.get('summary', {}).get('changed_files', 0)}",
        "",
        "## Documentation Targets",
        "",
    ]

    suggestions = _suggest_docs(records)
    if suggestions:
        for suggestion in suggestions:
            lines.append(f"- Review or update {suggestion}")
    else:
        lines.append("- No targeted documentation candidates were inferred from the current scope.")

    lines.extend(["", "## Draft Notes", ""])
    if grouped:
        for area_name in AREA_ORDER:
            area_records = grouped.get(area_name, [])
            if not area_records:
                continue
            lines.append(f"- {_draft_bullet_for_area(area_name, area_records, symbol_records=area_symbols.get(area_name, []))}")
    else:
        lines.append("- No tracked files were selected for the documentation draft.")

    lines.extend(["", "## Touched Symbols", ""])
    if touched_symbols:
        for path_text in sorted(touched_symbols):
            symbol_summary = _summarize_symbol_labels(touched_symbols.get(path_text, []), limit=5)
            if symbol_summary:
                lines.append(f"- {path_text}: {symbol_summary}")
    else:
        lines.append("- No symbol-level context was inferred from the current scope.")

    lines.extend(["", "## Recent Commit Context", ""])
    if commit_subjects:
        for subject_text in commit_subjects:
            lines.append(f"- {subject_text}")
    else:
        lines.append("- No recent commit subjects were available.")

    lines.extend(["", "## Files Considered", ""])
    if records:
        for record in records:
            lines.append(f"- [{record.get('area', 'root')}] {record.get('path', '(unknown)')}")
    else:
        lines.append("- No files selected.")
    lines.append("")
    return "\n".join(lines)


def generate_changelog_draft(workspace, version_text=None, release_date=None, changed_only=True):
    records = _changed_file_records(workspace, changed_only=changed_only)
    grouped = _records_grouped_by_area(records)
    touched_symbols = _collect_touched_symbols(workspace, records)
    area_symbols = _collect_area_symbol_summary(records, touched_symbols)
    commit_subjects = _recent_commit_subjects(workspace, limit=4)
    version_label = version_text or "Unreleased"
    date_label = release_date or _today_text()
    lines = [
        f"## [{version_label}] - {date_label}",
        "",
        "### Changed",
        "",
    ]

    if grouped:
        for area_name in AREA_ORDER:
            area_records = grouped.get(area_name, [])
            if not area_records:
                continue
            lines.append(f"- {_draft_bullet_for_area(area_name, area_records, symbol_records=area_symbols.get(area_name, []))}")
    else:
        lines.append("- No tracked changes were available for a changelog draft.")

    lines.extend(["", "### Notes", "", f"- Branch at draft time: {workspace.snapshot.get('git', {}).get('branch', 'unknown')}"])
    lines.append(f"- Snapshot generated at: {workspace.snapshot.get('generated_at', _utc_now_text())}")
    lines.append(f"- Files considered: {len(records)}")
    if commit_subjects:
        lines.append(f"- Recent commit context: {' | '.join(commit_subjects)}")
    lines.append("")
    return "\n".join(lines)


def _build_ai_context(workspace, task, changed_only=True):
    records = _changed_file_records(workspace, changed_only=changed_only)
    lines = [
        "# Project Librarian AI Context",
        "",
        f"Task: {task}",
        f"Generated: {workspace.snapshot.get('generated_at', _utc_now_text())}",
        f"Branch: {workspace.snapshot.get('git', {}).get('branch', 'unknown')}",
        "",
        "## Workspace Summary",
        "",
        format_workspace_stats(workspace),
        "",
        "## Current Changes",
        "",
        format_change_report(workspace.snapshot, limit=50, include_commits=True),
        "",
        "## Files In Scope",
        "",
    ]
    if records:
        for record in records:
            lines.append(f"- [{record.get('area', 'root')}] {record.get('path', '(unknown)')} | {record.get('title', '')}")
    else:
        lines.append("- No files selected.")
    lines.append("")
    return "\n".join(lines)


def run_ai_summary(workspace, task, mode="analysis", model=DEFAULT_AI_MODEL, changed_only=True, ollama_host=None):
    ai_status = collect_ai_runtime_status(workspace.repo_root, preferred_model=model, ollama_host=ollama_host)
    script_path = Path(ai_status.get("delegate_script"))
    if not ai_status.get("delegate_script_exists"):
        raise ProjectLibrarianError(f"AI delegate script not found at {script_path}")
    if not ai_status.get("delegate_script_executable"):
        raise ProjectLibrarianError(f"AI delegate script is not executable: {script_path}")
    if not ai_status.get("ollama_path"):
        raise ProjectLibrarianError("ollama is not installed or is not on PATH.")
    if not ai_status.get("ollama_reachable"):
        issue_text = ai_status.get("issues", ["Unable to query ollama."])[0]
        raise ProjectLibrarianError(
            f"Ollama is not reachable at {ai_status.get('ollama_host')}. {issue_text} Run 'project_librarian.py ai-doctor' for full diagnostics."
        )
    resolved_model = _resolve_ai_model(ai_status, model)

    context_text = _build_ai_context(workspace, task=task, changed_only=changed_only)
    context_path = _write_generated_output(
        workspace.output_dir,
        AI_CONTEXT_DIR_NAME,
        "ai_context",
        context_text,
    )

    command = [
        "bash",
        str(script_path),
        "--task",
        task,
        "--mode",
        mode,
        "--context",
        str(context_path),
        "--model",
        resolved_model,
    ]
    env = os.environ.copy()
    if ollama_host:
        env["OLLAMA_HOST"] = str(ollama_host)

    completed = subprocess.run(
        command,
        cwd=workspace.repo_root,
        check=False,
        capture_output=True,
        text=True,
        env=env,
    )
    if completed.returncode != 0:
        error_text = (completed.stderr or completed.stdout or "AI summary failed.").strip()
        raise ProjectLibrarianError(error_text)

    report_path = None
    for line in completed.stdout.splitlines():
        stripped = line.strip()
        if stripped.startswith("Saved:"):
            report_path = stripped.split(":", 1)[1].strip()
            break
    return {
        "context_path": context_path,
        "report_path": report_path,
        "model": resolved_model,
        "stdout": completed.stdout.strip(),
    }


def run_repl(repo_root=None, output_dir=None, refresh_first=False):
    workspace = LibrarianWorkspace.load(
        repo_root=repo_root,
        output_dir=output_dir,
        refresh_if_missing=True,
        refresh_first=refresh_first,
    )

    print(_format_repl_welcome(workspace))
    while True:
        try:
            raw_command = input("librarian> ").strip()
        except EOFError:
            print()
            return 0
        if not raw_command:
            continue
        try:
            parts = shlex.split(raw_command)
        except ValueError as exc:
            print(f"Command parse error: {exc}")
            continue
        if not parts:
            continue
        command_name = parts[0].lower()
        command_args = parts[1:]

        if command_name in {"quit", "exit"}:
            return 0
        if command_name == "help":
            print(
                "Commands: search <query> | symbols <query> | files <query> | changes | history | show <path> | docs-draft | changelog-draft | ai-models | ai-doctor | refresh | stats | quit"
            )
            continue
        if command_name == "refresh":
            result = workspace.refresh()
            snapshot = result["snapshot"]
            print(
                f"Refreshed snapshot: {snapshot['summary']['files']} files, "
                f"{snapshot['summary']['symbols']} symbols, "
                f"{snapshot['summary']['changed_files']} changed files."
            )
            continue
        if command_name == "stats":
            print(format_workspace_stats(workspace))
            continue
        if command_name == "changes":
            print(format_change_report(workspace.snapshot))
            continue
        if command_name == "history":
            print(format_history_report(workspace.history))
            continue
        if command_name == "show" and command_args:
            print(show_file_excerpt(workspace, command_args[0]))
            continue
        if command_name == "docs-draft":
            draft_path = _write_generated_output(
                workspace.output_dir,
                DRAFTS_DIR_NAME,
                "docs_draft",
                generate_docs_draft(workspace),
            )
            print(f"Saved docs draft: {draft_path}")
            continue
        if command_name == "changelog-draft":
            draft_path = _write_generated_output(
                workspace.output_dir,
                DRAFTS_DIR_NAME,
                "changelog_draft",
                generate_changelog_draft(workspace),
            )
            print(f"Saved changelog draft: {draft_path}")
            continue
        if command_name == "ai-models":
            print(format_ai_model_list(collect_ai_runtime_status(workspace.repo_root)))
            continue
        if command_name == "ai-doctor":
            print(format_ai_status_report(collect_ai_runtime_status(workspace.repo_root)))
            continue
        if command_name in {"search", "symbols", "files"} and command_args:
            query = " ".join(command_args)
            scope = "all" if command_name == "search" else command_name
            print(format_search_results(search_snapshot(workspace.snapshot, workspace.corpus, query, scope=scope)))
            continue
        print("Unknown command. Type 'help' for usage.")


def parse_args():
    parser = argparse.ArgumentParser(description="Refresh and query a local in-memory project librarian.")
    parser.add_argument(
        "--repo-root",
        default=str(_repo_root_from_here()),
        help="Repository root to catalog. Defaults to the current repository.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help=f"Directory that receives {SNAPSHOT_NAME}, {CORPUS_NAME}, and {HISTORY_NAME}.",
    )

    subparsers = parser.add_subparsers(dest="command")

    subparsers.add_parser("refresh", help="Refresh the librarian snapshot, corpus, and change history.")

    search_parser = subparsers.add_parser("search", help="Search files and symbols from the current snapshot.")
    search_parser.add_argument("query", help="Query string to search for.")
    search_parser.add_argument("--scope", choices=("all", "files", "symbols"), default="all")
    search_parser.add_argument("--limit", type=int, default=20)
    search_parser.add_argument("--area", choices=AREA_ORDER)
    search_parser.add_argument("--path", dest="path_filter")
    search_parser.add_argument("--changed-only", action="store_true")

    changes_parser = subparsers.add_parser("changes", help="Show changed files recorded in the current snapshot.")
    changes_parser.add_argument("--limit", type=int, default=20)
    changes_parser.add_argument("--status")
    changes_parser.add_argument("--area", choices=AREA_ORDER)
    changes_parser.add_argument("--path", dest="path_filter")
    changes_parser.add_argument("--no-commits", action="store_true", help="Hide recent commit context.")

    history_parser = subparsers.add_parser("history", help="Show recent librarian refresh history.")
    history_parser.add_argument("--limit", type=int, default=10)

    show_parser = subparsers.add_parser("show", help="Show a file excerpt from the RAM-loaded corpus.")
    show_parser.add_argument("path", help="Indexed file path, suffix, or unique substring.")
    show_parser.add_argument("--query", help="Anchor the excerpt around the first matching line.")
    show_parser.add_argument("--line", type=int, help="1-based line to center the excerpt on.")
    show_parser.add_argument("--context", type=int, default=3)

    docs_parser = subparsers.add_parser("docs-draft", help="Generate a documentation update draft from the current snapshot.")
    docs_parser.add_argument("--title")
    docs_parser.add_argument("--all-files", action="store_true", help="Draft from the full indexed workspace instead of only changed files.")
    docs_parser.add_argument("--output")
    docs_parser.add_argument("--apply", action="store_true", help="Write the generated documentation block into a target markdown file.")
    docs_parser.add_argument("--target", default=README_TARGET_NAME, help="Markdown file to update when --apply is used.")

    changelog_parser = subparsers.add_parser("changelog-draft", help="Generate a changelog draft from the current snapshot.")
    changelog_parser.add_argument("--version")
    changelog_parser.add_argument("--date")
    changelog_parser.add_argument("--all-files", action="store_true", help="Draft from the full indexed workspace instead of only changed files.")
    changelog_parser.add_argument("--output")
    changelog_parser.add_argument("--apply", action="store_true", help="Write or replace the release entry in a changelog file.")
    changelog_parser.add_argument("--target", default=CHANGELOG_TARGET_NAME, help="Changelog file to update when --apply is used.")

    ai_parser = subparsers.add_parser("ai-summary", help="Ask the local AI helper to summarize the current workspace and changes.")
    ai_parser.add_argument("--task", default="Summarize the current repository changes and likely next actions.")
    ai_parser.add_argument("--mode", choices=("analysis", "review"), default="analysis")
    ai_parser.add_argument("--model", default=DEFAULT_AI_MODEL)
    ai_parser.add_argument("--all-files", action="store_true", help="Include the full indexed workspace context instead of only changed files.")
    ai_parser.add_argument("--ollama-host", help="Optional OLLAMA_HOST override for local AI calls.")

    ai_models_parser = subparsers.add_parser("ai-models", help="List locally available Ollama models for the selected host.")
    ai_models_parser.add_argument("--model", default=DEFAULT_AI_MODEL, help="Preferred model to compare against the local list.")
    ai_models_parser.add_argument("--ollama-host", help="Optional OLLAMA_HOST override for local AI calls.")

    ai_doctor_parser = subparsers.add_parser("ai-doctor", help="Show Ollama and delegate-script diagnostics for local AI usage.")
    ai_doctor_parser.add_argument("--model", default=DEFAULT_AI_MODEL, help="Preferred model to validate against the local list.")
    ai_doctor_parser.add_argument("--ollama-host", help="Optional OLLAMA_HOST override for local AI calls.")

    mcp_parser = subparsers.add_parser("mcp-server", help="Run the librarian as a persistent MCP server with a shared in-memory workspace.")
    mcp_parser.add_argument("--transport", choices=("stdio", "sse", "streamable-http"), default=DEFAULT_MCP_TRANSPORT)
    mcp_parser.add_argument("--host", default=DEFAULT_MCP_HOST, help="Host to bind for HTTP-based transports.")
    mcp_parser.add_argument("--port", type=int, default=DEFAULT_MCP_PORT, help="Port to bind for HTTP-based transports.")
    mcp_parser.add_argument("--refresh", action="store_true", help="Refresh the workspace before starting the MCP server.")
    mcp_parser.add_argument(
        "--refresh-interval",
        type=float,
        default=DEFAULT_REFRESH_INTERVAL_SECONDS,
        help="Background refresh interval in seconds for the second worker thread. Use 0 to disable it.",
    )
    mcp_parser.add_argument(
        "--auth-token",
        help=(
            "Optional shared token for the browser dashboard and HTTP MCP endpoint. "
            f"If omitted, {HTTP_AUTH_ENV_NAME} is used when present. Use 'auto' to generate a token for this run."
        ),
    )
    mcp_parser.add_argument("--log-level", choices=("DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"), default="INFO")

    repl_parser = subparsers.add_parser("repl", help="Load the project into RAM and interactively search it.")
    repl_parser.add_argument("--refresh", action="store_true", help="Refresh the snapshot before entering the REPL.")

    subparsers.add_parser("stats", help="Show summary information for the current snapshot.")
    return parser.parse_args()


def main():
    args = parse_args()
    repo_root = Path(args.repo_root).resolve()
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = (repo_root / output_dir).resolve()

    if not args.command:
        return run_repl(repo_root=repo_root, output_dir=output_dir, refresh_first=False)

    if args.command == "refresh":
        result = build_librarian_snapshot(repo_root=repo_root, output_dir=output_dir)
        summary = result["snapshot"]["summary"]
        print(
            f"Refreshed project librarian: {summary['files']} files, "
            f"{summary['symbols']} symbols, {summary['changed_files']} changed files."
        )
        print(f"Snapshot: {result['snapshot_path']}")
        print(f"History: {result['history_path']}")
        print(f"Corpus: {result['corpus_path']}")
        return 0

    workspace = LibrarianWorkspace.load(repo_root=repo_root, output_dir=output_dir, refresh_if_missing=True)

    if args.command == "search":
        print(
            format_search_results(
                search_snapshot(
                    workspace.snapshot,
                    workspace.corpus,
                    args.query,
                    scope=args.scope,
                    limit=args.limit,
                    area=args.area,
                    changed_only=args.changed_only,
                    path_filter=args.path_filter,
                )
            )
        )
        return 0

    if args.command == "changes":
        print(
            format_change_report(
                workspace.snapshot,
                limit=args.limit,
                status_filter=args.status,
                area=args.area,
                path_filter=args.path_filter,
                include_commits=not args.no_commits,
            )
        )
        return 0

    if args.command == "history":
        print(format_history_report(workspace.history, limit=args.limit))
        return 0

    if args.command == "show":
        print(show_file_excerpt(workspace, args.path, query=args.query, line=args.line, context=args.context))
        return 0

    if args.command == "docs-draft":
        draft_content = generate_docs_draft(workspace, title=args.title, changed_only=not args.all_files)
        if args.apply:
            target_path = apply_docs_update(repo_root, draft_content, target_path=args.target)
            print(f"Updated docs target: {target_path}")
        else:
            draft_path = _write_generated_output(
                workspace.output_dir,
                DRAFTS_DIR_NAME,
                "docs_draft",
                draft_content,
                output_path=args.output,
                output_base_dir=repo_root,
            )
            print(f"Saved docs draft: {draft_path}")
        return 0

    if args.command == "changelog-draft":
        draft_content = generate_changelog_draft(
            workspace,
            version_text=args.version,
            release_date=args.date,
            changed_only=not args.all_files,
        )
        if args.apply:
            version_label = args.version or "Unreleased"
            target_path = apply_changelog_update(repo_root, draft_content, version_label=version_label, target_path=args.target)
            print(f"Updated changelog target: {target_path}")
        else:
            draft_path = _write_generated_output(
                workspace.output_dir,
                DRAFTS_DIR_NAME,
                "changelog_draft",
                draft_content,
                output_path=args.output,
                output_base_dir=repo_root,
            )
            print(f"Saved changelog draft: {draft_path}")
        return 0

    if args.command == "ai-summary":
        result = run_ai_summary(
            workspace,
            task=args.task,
            mode=args.mode,
            model=args.model,
            changed_only=not args.all_files,
            ollama_host=args.ollama_host,
        )
        print(f"AI context: {result['context_path']}")
        print(f"AI model: {result['model']}")
        if result.get("report_path"):
            print(f"AI report: {result['report_path']}")
        else:
            print(result.get("stdout") or "AI summary completed without a saved report path.")
        return 0

    if args.command == "ai-models":
        print(format_ai_model_list(collect_ai_runtime_status(repo_root, preferred_model=args.model, ollama_host=args.ollama_host)))
        return 0

    if args.command == "ai-doctor":
        print(format_ai_status_report(collect_ai_runtime_status(repo_root, preferred_model=args.model, ollama_host=args.ollama_host)))
        return 0

    if args.command == "mcp-server":
        return run_librarian_mcp_server(
            repo_root=repo_root,
            output_dir=output_dir,
            transport=args.transport,
            host=args.host,
            port=args.port,
            refresh_first=args.refresh,
            refresh_interval_seconds=args.refresh_interval,
            log_level=args.log_level,
            auth_token=args.auth_token,
        )

    if args.command == "stats":
        print(format_workspace_stats(workspace))
        return 0

    if args.command == "repl":
        return run_repl(repo_root=repo_root, output_dir=output_dir, refresh_first=args.refresh)

    raise ProjectLibrarianError(f"Unsupported command: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())